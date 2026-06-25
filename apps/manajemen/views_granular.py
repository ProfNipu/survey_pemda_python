"""
Granular Permission Management Views
For managing Functions, Controls, Modules, Rules, and Role-Rules
"""

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.models import Group
from django.contrib.auth import get_user_model
from django.contrib.auth.decorators import login_required
from django.contrib.auth import update_session_auth_hash
from django.contrib import messages
from django.views.decorators.csrf import ensure_csrf_cookie
from django.db.models import Count, Q
from django.http import JsonResponse, HttpResponse
from django_tables2 import RequestConfig
import csv
from io import BytesIO
from django.shortcuts import redirect
from django.contrib import messages
from functools import wraps
from django.conf import settings

User = get_user_model()

from .models import (
    PermissionFunction,
    PermissionControl,
    PermissionModule,
    PermissionRule,
    RoleRule,
    UserTableSelection,
    MenuItem,
    MenuCategory,
)
from .tables import UserTable, RoleTable, FunctionTable, MenuItemTable, MenuCategoryTable, ControlTable, ModuleManageTable
from .forms import MenuItemForm, MenuCategoryForm
from .helpers import is_superadmin
from .decorators import permission_required, permission_required_403


def staff_or_superadmin_required(view_func):
    @wraps(view_func)
    def _wrapped(request, *args, **kwargs):
        user = getattr(request, 'user', None)
        if user and user.is_authenticated and is_superadmin(user):
            return view_func(request, *args, **kwargs)
        if not user or not user.is_authenticated:
            return redirect('login')
        messages.warning(request, 'Akses halaman ini memerlukan superadmin.')
        return redirect('manajemen_aplikasi:dashboard')
    return _wrapped


# ==================== EXPORT HELPERS ====================

def export_users_csv(users):
    """Export users to CSV"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="users_export.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['No', 'Username', 'Email', 'Name', 'Roles', 'Status', 'Staff', 'Superuser'])
    
    for i, user in enumerate(users, 1):
        roles = ', '.join([g.name for g in user.groups.all()])
        status = 'Active' if user.is_active else 'Inactive'
        writer.writerow([
            i,
            user.username,
            user.email or '-',
            user.name or '-',
            roles or '-',
            status,
            'No',
            'No'
        ])
    
    return response


def export_users_excel(users):
    """Export users to Excel (using openpyxl if available, else CSV)"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Users Export"
        
        # Header styling
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        # Headers
        headers = ['No', 'Username', 'Email', 'Name', 'Roles', 'Status', 'Staff', 'Superuser']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Data
        for i, user in enumerate(users, 2):
            roles = ', '.join([g.name for g in user.groups.all()])
            status = 'Active' if user.is_active else 'Inactive'
            
            ws.cell(row=i, column=1, value=i-1)
            ws.cell(row=i, column=2, value=user.username)
            ws.cell(row=i, column=3, value=user.email or '-')
            ws.cell(row=i, column=4, value=user.name or '-')
            ws.cell(row=i, column=5, value=roles or '-')
            ws.cell(row=i, column=6, value=status)
            ws.cell(row=i, column=7, value='No')
            ws.cell(row=i, column=8, value='No')
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 25
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 10
        ws.column_dimensions['H'].width = 12
        
        # Save to response
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="users_export.xlsx"'
        return response
        
    except ImportError:
        # Fallback to CSV if openpyxl not installed
        return export_users_csv(users)


def export_users_pdf(users):
    """Export users to PDF with professional styling (A4 portrait)"""
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER
        
        output = BytesIO()
        
        # A4 portrait dengan margins
        doc = SimpleDocTemplate(
            output, 
            pagesize=A4,
            leftMargin=1*cm,
            rightMargin=1*cm,
            topMargin=1.5*cm,
            bottomMargin=1.5*cm
        )
        elements = []
        
        # Title - "Users Export"
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.black,
            spaceAfter=20,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        )
        title = Paragraph("<b>Users Export</b>", title_style)
        elements.append(title)
        elements.append(Spacer(1, 0.3*cm))
        
        # Table data - sesuai gambar
        # Style untuk cell content (word wrap)
        cell_style = ParagraphStyle(
            'CellStyle',
            parent=styles['Normal'],
            fontSize=8,
            leading=10,
            wordWrap='CJK'
        )
        
        # Header row
        data = [['Username', 'Email', 'Name', 'Roles', 'Status']]
        
        # Data rows dengan Paragraph untuk word wrap
        for user in users:
            roles = ', '.join([g.name for g in user.groups.all()])
            status = 'Active' if user.is_active else 'Inactive'
            
            data.append([
                Paragraph(user.username, cell_style),
                Paragraph(user.email or '-', cell_style),
                Paragraph(user.name or '-', cell_style),
                Paragraph(roles or '-', cell_style),
                Paragraph(status, cell_style)
            ])
        
        # Create table dengan styling seperti gambar
        # Column widths: username, email, name, roles, status
        col_widths = [3*cm, 4*cm, 5*cm, 4*cm, 2*cm]
        
        table = Table(data, colWidths=col_widths, repeatRows=1)
        
        # Styling - header abu-abu seperti gambar
        table.setStyle(TableStyle([
            # Header row (abu-abu gelap)
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#666666')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('TOPPADDING', (0, 0), (-1, 0), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            
            # Data rows
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 1), (-1, -1), 'TOP'),  # Align text to top for word wrap
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('TOPPADDING', (0, 1), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 5),
            ('LEFTPADDING', (0, 0), (-1, -1), 5),
            ('RIGHTPADDING', (0, 0), (-1, -1), 5),
            
            # Grid
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('BOX', (0, 0), (-1, -1), 1, colors.black),
            
            # Alternating row colors (zebra striping)
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f5f5')]),
        ]))
        
        elements.append(table)
        
        # Build PDF
        doc.build(elements)
        output.seek(0)
        
        response = HttpResponse(output.read(), content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="users_export.pdf"'
        return response
        
    except ImportError:
        # Fallback to CSV if reportlab not installed
        return export_users_csv(users)


# ==================== ROLE EXPORT HELPERS ====================

def export_roles_csv(roles):
    """Export roles to CSV"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="roles_export.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['No', 'ID', 'Nama Role', 'Users', 'Permissions'])
    
    for i, role in enumerate(roles, 1):
        users_count = role.user_set.count()
        permissions_count = role.permissions.count()
        writer.writerow([
            i,
            role.id,
            role.name,
            users_count,
            permissions_count
        ])
    
    return response


def export_roles_excel(roles):
    """Export roles to Excel (using openpyxl if available, else CSV)"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Roles Export"
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Headers
        headers = ['No', 'ID', 'Nama Role', 'Users', 'Permissions']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        
        # Data rows
        for row, role in enumerate(roles, 2):
            users_count = role.user_set.count()
            permissions_count = role.permissions.count()
            
            data = [
                row-1,  # No
                role.id,
                role.name,
                users_count,
                permissions_count
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = border
                if col == 1 or col == 2 or col == 4 or col == 5:  # Center align numbers
                    cell.alignment = Alignment(horizontal="center")
        
        # Auto-adjust column widths
        column_widths = [8, 8, 30, 10, 15]  # Custom widths
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
        
        # Save to response
        from io import BytesIO
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="roles_export.xlsx"'
        return response
        
    except ImportError:
        # Fallback to CSV if openpyxl not installed
        return export_roles_csv(roles)


# ==================== FUNCTION EXPORT HELPERS ====================

def export_functions_csv(functions):
    """Export functions to CSV"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="functions_export.csv"'
    writer = csv.writer(response)
    writer.writerow(['No', 'ID', 'Nama Fungsi', 'Label', 'Deskripsi', 'Rules'])
    for i, f in enumerate(functions, 1):
        writer.writerow([
            i,
            f.id,
            f.nama_fungsi,
            f.label_fungsi,
            (f.deskripsi_fungsi or '')[:200],
            getattr(f, 'rule_count', 0)
        ])
    return response


def export_functions_excel(functions):
    """Export functions to Excel (openpyxl if available, else CSV)"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Functions Export"
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        headers = ['No', 'ID', 'Nama Fungsi', 'Label', 'Deskripsi', 'Rules']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        
        for row, f in enumerate(functions, 2):
            data = [
                row-1,
                f.id,
                f.nama_fungsi,
                f.label_fungsi,
                (f.deskripsi_fungsi or ''),
                getattr(f, 'rule_count', 0)
            ]
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = border
                if col in (1, 2, 6):
                    cell.alignment = Alignment(horizontal="center")
        
        # Column widths
        widths = [6, 8, 24, 24, 40, 10]
        for col, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="functions_export.xlsx"'
        return response
    except ImportError:
        return export_functions_csv(functions)

# ==================== CONTROL EXPORT HELPERS ====================

def export_controls_csv(controls):
    """Export controls to CSV"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="controls_export.csv"'
    writer = csv.writer(response)
    writer.writerow(['No', 'ID', 'Nama Kontrol', 'Label', 'Deskripsi', 'Rules'])
    for i, c in enumerate(controls, 1):
        writer.writerow([
            i,
            c.id,
            c.nama_kontrol,
            c.label_kontrol,
            (c.deskripsi_kontrol or '')[:200],
            getattr(c, 'rule_count', 0)
        ])
    return response


def export_controls_excel(controls):
    """Export controls to Excel (openpyxl if available, else CSV)"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        wb = Workbook()
        ws = wb.active
        ws.title = "Controls Export"
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        headers = ['No', 'ID', 'Nama Kontrol', 'Label', 'Deskripsi', 'Rules']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        for row, c in enumerate(controls, 2):
            data = [row-1, c.id, c.nama_kontrol, c.label_kontrol, (c.deskripsi_kontrol or ''), getattr(c, 'rule_count', 0)]
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = border
                if col in (1, 2, 6):
                    cell.alignment = Alignment(horizontal="center")
        widths = [6, 8, 24, 24, 40, 10]
        for col, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="controls_export.xlsx"'
        return response
    except ImportError:
        return export_controls_csv(controls)


def export_controls_pdf(controls):
    """Export controls to PDF (A4 portrait)"""
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=18)
        elements = []
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle('Title', parent=styles['Heading1'], alignment=TA_CENTER, fontSize=16)
        elements.append(Paragraph('CONTROLS EXPORT REPORT', title_style))
        elements.append(Spacer(1, 12))
        data = [['No', 'ID', 'Nama Kontrol', 'Label', 'Rules']]
        for i, c in enumerate(controls, 1):
            data.append([str(i), str(c.id), c.nama_kontrol, c.label_kontrol, str(getattr(c, 'rule_count', 0))])
        table = Table(data, colWidths=[0.6*inch, 0.6*inch, 2.6*inch, 2.2*inch, 0.8*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ALIGN', (0, 1), (1, -1), 'CENTER'),
            ('ALIGN', (4, 1), (4, -1), 'CENTER'),
        ]))
        elements.append(table)
        doc.build(elements)
        buffer.seek(0)
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="controls_export.pdf"'
        return response
    except ImportError:
        return export_controls_csv(controls)


def export_modules_csv(modules):
    """Export modules to CSV"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="modules_export.csv"'
    writer = csv.writer(response)
    writer.writerow(['No', 'Order', 'Nama Module', 'Label', 'Icon', 'Status', 'Rules'])
    for i, m in enumerate(modules, 1):
        writer.writerow([
            i,
            m.order,
            m.nama_module,
            m.label_module,
            m.icon,
            'Active' if m.is_active else 'Inactive',
            getattr(m, 'rule_count', 0)
        ])
    return response


def export_modules_excel(modules):
    """Export modules to Excel (openpyxl if available, else CSV)"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        wb = Workbook()
        ws = wb.active
        ws.title = "Modules Export"
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        headers = ['No', 'Order', 'Nama Module', 'Label', 'Icon', 'Status', 'Rules']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        for row, m in enumerate(modules, 2):
            data = [row-1, m.order, m.nama_module, m.label_module, m.icon, 'Active' if m.is_active else 'Inactive', getattr(m, 'rule_count', 0)]
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = border
                if col in (1, 2, 7):
                    cell.alignment = Alignment(horizontal="center")
        widths = [6, 8, 24, 24, 20, 12, 10]
        for col, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="modules_export.xlsx"'
        return response
    except ImportError:
        return export_modules_csv(modules)


def export_modules_pdf(modules):
    """Export modules to PDF (A4 portrait)"""
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=18)
        elements = []
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle('Title', parent=styles['Heading1'], alignment=TA_CENTER, fontSize=16)
        elements.append(Paragraph('MODULES EXPORT REPORT', title_style))
        elements.append(Spacer(1, 12))
        data = [['No', 'Order', 'Label', 'Nama', 'Rules']]
        for i, m in enumerate(modules, 1):
            data.append([str(i), str(m.order), m.label_module, m.nama_module, str(getattr(m, 'rule_count', 0))])
        table = Table(data, colWidths=[0.6*inch, 0.6*inch, 2.4*inch, 2.4*inch, 0.8*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ALIGN', (0, 1), (1, -1), 'CENTER'),
            ('ALIGN', (4, 1), (4, -1), 'CENTER'),
        ]))
        elements.append(table)
        doc.build(elements)
        buffer.seek(0)
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="modules_export.pdf"'
        return response
    except ImportError:
        return export_modules_csv(modules)

def export_functions_pdf(functions):
    """Export functions to PDF (A4 portrait)"""
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=18)
        elements = []
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle('Title', parent=styles['Heading1'], alignment=TA_CENTER, fontSize=16)
        elements.append(Paragraph('FUNCTIONS EXPORT REPORT', title_style))
        elements.append(Spacer(1, 12))
        
        data = [['No', 'ID', 'Nama Fungsi', 'Label', 'Rules']]
        for i, f in enumerate(functions, 1):
            data.append([
                str(i),
                str(f.id),
                f.nama_fungsi,
                f.label_fungsi,
                str(getattr(f, 'rule_count', 0))
            ])
        
        table = Table(data, colWidths=[0.6*inch, 0.6*inch, 2.6*inch, 2.2*inch, 0.8*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ALIGN', (0, 1), (1, -1), 'CENTER'),
            ('ALIGN', (4, 1), (4, -1), 'CENTER'),
        ]))
        elements.append(table)
        
        doc.build(elements)
        buffer.seek(0)
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="functions_export.pdf"'
        return response
    except ImportError:
        return export_functions_csv(functions)

def export_roles_pdf(roles):
    """Export roles to PDF with professional styling (A4 portrait)"""
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
        from datetime import datetime
        from io import BytesIO
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=18)
        
        # Container for the 'Flowable' objects
        elements = []
        
        # Define styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=TA_CENTER,
            textColor=colors.HexColor('#2563eb')
        )
        
        # Title
        title = Paragraph("ROLES EXPORT REPORT", title_style)
        elements.append(title)
        
        # Export info
        export_info_style = ParagraphStyle(
            'ExportInfo',
            parent=styles['Normal'],
            fontSize=10,
            spaceAfter=20,
            alignment=TA_CENTER,
            textColor=colors.grey
        )
        
        export_time = datetime.now().strftime('%d %B %Y, %H:%M:%S')
        export_info = Paragraph(f"Generated on: {export_time} | Total Roles: {len(roles)}", export_info_style)
        elements.append(export_info)
        
        # Prepare table data
        data = [['No', 'ID', 'Nama Role', 'Users', 'Permissions']]
        
        for i, role in enumerate(roles, 1):
            users_count = role.user_set.count()
            permissions_count = role.permissions.count()
            
            data.append([
                str(i),
                str(role.id),
                str(role.name),
                str(users_count),
                str(permissions_count)
            ])
        
        # Create table
        table = Table(data, colWidths=[0.8*inch, 0.8*inch, 3*inch, 1*inch, 1.2*inch])
        
        # Table styling
        table.setStyle(TableStyle([
            # Header styling
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            
            # Data rows styling
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('ALIGN', (0, 1), (1, -1), 'CENTER'),  # No, ID columns center
            ('ALIGN', (2, 1), (2, -1), 'LEFT'),    # Name column left
            ('ALIGN', (3, 1), (-1, -1), 'CENTER'), # Users, Permissions center
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
            
            # Grid styling
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        
        elements.append(table)
        
        # Footer
        elements.append(Spacer(1, 20))
        footer_style = ParagraphStyle(
            'Footer',
            parent=styles['Normal'],
            fontSize=8,
            alignment=TA_CENTER,
            textColor=colors.grey
        )
        footer = Paragraph(f"This report was generated automatically by {getattr(settings, 'APP_NAME', 'aplikasi-test')}", footer_style)
        elements.append(footer)
        
        # Build PDF
        doc.build(elements)
        
        buffer.seek(0)
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="roles_export.pdf"'
        return response
        
    except ImportError:
        # Fallback to CSV if reportlab not installed
        return export_roles_csv(roles)


# ==================== DASHBOARD ====================

@ensure_csrf_cookie
@permission_required_403('pengaturan', 'permission_dashboard', 'view')
def granular_dashboard(request):
    """Main dashboard for granular permission management + Modules datatable reusable"""
    
    # Get statistics
    stats = {
        'functions': PermissionFunction.objects.count(),
        'controls': PermissionControl.objects.count(),
        'modules': PermissionModule.objects.filter(is_active=True).count(),
        'rules': PermissionRule.objects.filter(is_active=True).count(),
        'roles': Group.objects.count(),
        'users': User.objects.filter(is_active=True).count(),
    }
    
    # Recent items
    recent_functions = PermissionFunction.objects.order_by('-created_at')[:5]
    recent_controls = PermissionControl.objects.order_by('-created_at')[:5]
    recent_modules = PermissionModule.objects.order_by('-created_at')[:5]

    # ===== MODULES DATATABLE (Reusable) =====
    modules_qs = PermissionModule.objects.annotate(
        rule_count=Count('rules')
    )

    # EXPORT ALL (POST)
    if request.method == 'POST' and 'export_all' in request.POST:
        export_format = request.POST.get('export_all')
        search_query = request.POST.get('search', '').strip()
        all_modules = modules_qs
        if search_query:
            all_modules = all_modules.filter(
                Q(label_module__icontains=search_query) |
                Q(nama_module__icontains=search_query)
            )
        all_modules = all_modules.order_by('order', 'nama_module')
        if export_format == 'excel':
            try:
                return export_modules_excel(all_modules)
            except Exception:
                return export_modules_csv(all_modules)
        elif export_format == 'pdf':
            try:
                return export_modules_pdf(all_modules)
            except Exception:
                return export_modules_csv(all_modules)

    # AJAX save/load selection for modules dashboard
    if (
        request.method == 'POST'
        and request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        and (request.headers.get('Content-Type') or '').startswith('application/json')
    ):
        import json
        data = json.loads(request.body)
        action = data.get('action')
        if action == 'save_selection':
            page_key = data.get('page_key', 'modules_dashboard')
            selected_ids = data.get('selected_ids', [])
            UserTableSelection.objects.update_or_create(
                user=request.user,
                page_key=page_key,
                defaults={'selected_ids': selected_ids}
            )
            return JsonResponse({'success': True, 'count': len(selected_ids)})
        elif action == 'load_selection':
            page_key = data.get('page_key', 'modules_dashboard')
            try:
                sel = UserTableSelection.objects.get(user=request.user, page_key=page_key)
                return JsonResponse({'success': True, 'selected_ids': sel.selected_ids})
            except UserTableSelection.DoesNotExist:
                return JsonResponse({'success': True, 'selected_ids': []})

    # BULK ACTIONS (exports only for dashboard)
    if request.method == 'POST' and 'action' in request.POST:
        action = request.POST.get('action')
        selected_ids = request.POST.getlist('selected_ids')
        if selected_ids:
            sel_modules = modules_qs.filter(id__in=selected_ids).order_by('order', 'nama_module')
            if action == 'export_csv':
                return export_modules_csv(sel_modules)
            elif action == 'export_excel':
                try:
                    return export_modules_excel(sel_modules)
                except Exception:
                    return export_modules_csv(sel_modules)
            elif action == 'export_pdf':
                try:
                    return export_modules_pdf(sel_modules)
                except Exception:
                    return export_modules_csv(sel_modules)
        else:
            messages.error(request, 'Tidak ada item yang dipilih')
            return redirect('manajemen_aplikasi:dashboard')

    # Regular GET for modules datatable
    search_query = request.GET.get('search', '').strip()
    modules = modules_qs
    if search_query:
        modules = modules.filter(
            Q(label_module__icontains=search_query) |
            Q(nama_module__icontains=search_query)
        )
    modules = modules.order_by('order', 'nama_module')

    from .tables import ModuleTable  # local import to avoid circular in top changes
    modules_table = ModuleTable(modules)
    per_page = request.GET.get('per_page', '10')
    try:
        per_page = int(per_page)
        if per_page not in [10, 25, 50, 100]:
            per_page = 10
    except (ValueError, TypeError):
        per_page = 10
    RequestConfig(request, paginate={'per_page': per_page}).configure(modules_table)

    # Context
    context = {
        'stats': stats,
        'recent_functions': recent_functions,
        'recent_controls': recent_controls,
        'recent_modules': recent_modules,
        'table': modules_table,
        'total': modules.count(),
        'search_query': search_query,
    }

    # HTMX partial (modules table only)
    is_htmx = request.headers.get('HX-Request') or request.headers.get('X-Requested-With') == 'XMLHttpRequest'
    if is_htmx:
        return render(request, 'manajemen_aplikasi/access/ma_ap_modules/partials/_table.html', context)

    return render(request, 'manajemen_aplikasi/access/dashboard.html', context)


 
 
# ==================== ROLE MANAGEMENT ====================


 
 
# ==================== ROLE MANAGEMENT ====================


# ==================== ROLE MANAGEMENT ====================

 


# ==================== USER MANAGEMENT ====================










# ==================== USER PROFILE & PASSWORD ====================

@login_required
def change_password(request):
    """Legacy shim: redirect to accounts:change_password (canonical)."""
    return redirect('accounts:change_password')


@login_required
def user_profile(request):
    """Legacy shim: redirect to accounts:profile (canonical)."""
    return redirect('accounts:profile')
