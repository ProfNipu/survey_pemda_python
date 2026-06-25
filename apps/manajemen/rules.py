from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.urls import reverse
import json
from django.db.models import Count, Q, Value
from django.db.models.functions import Concat
from django_tables2 import RequestConfig
from apps.manajemen.tables import RuleTable
from django.core.exceptions import PermissionDenied

from apps.manajemen.decorators import permission_required_403
from apps.manajemen.models import (
    PermissionFunction,
    PermissionControl,
    PermissionModule,
    PermissionRule,
    RoleRule,
    UserTableSelection,
)


@permission_required_403('pengaturan', 'permission_rule', 'view')
def rule_list(request):
    """List all permission rules with reusable datatable (consistent with other pages)"""

    def _require(func_name: str) -> None:
        try:
            from apps.manajemen.helpers import check_permission
            if not check_permission(request.user, 'pengaturan', 'permission_rule', func_name):
                raise PermissionDenied
        except PermissionDenied:
            raise
        except Exception:
            raise PermissionDenied

    # ---- EXPORT ALL (POST) ----
    if request.method == 'POST' and 'export_all' in request.POST:
        _require('export')
        export_format = (request.POST.get('export_all') or '').lower()
        search_query = (request.POST.get('search') or '').strip()
        module_filter = (request.POST.get('module') or '').strip()
        rules_all = PermissionRule.objects.select_related('module','control','function').annotate(
            role_count=Count('role_assignments'),
            perm_str=Concat('module__nama_module', Value('.'), 'control__nama_kontrol', Value('.'), 'function__nama_fungsi')
        )
        if search_query:
            terms = [t for t in search_query.split() if t]
            for t in terms:
                rules_all = rules_all.filter(
                    Q(module__label_module__icontains=t) |
                    Q(module__nama_module__icontains=t) |
                    Q(control__label_kontrol__icontains=t) |
                    Q(control__nama_kontrol__icontains=t) |
                    Q(function__label_fungsi__icontains=t) |
                    Q(function__nama_fungsi__icontains=t) |
                    Q(perm_str__icontains=t)
                )
        if module_filter:
            rules_all = rules_all.filter(module__nama_module=module_filter)
        rules_all = rules_all.order_by('module__order','control__nama_kontrol','function__nama_fungsi')
        if export_format == 'excel':
            try:
                return _export_rules_excel(rules_all)
            except Exception:
                return _export_rules_csv(rules_all)
        elif export_format == 'pdf':
            try:
                return _export_rules_pdf(rules_all)
            except Exception:
                return _export_rules_csv(rules_all)
        else:
            return _export_rules_csv(rules_all)
    # ---- AJAX SAVE/LOAD SELECTIONS ----
    if request.method == 'POST' and request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        content_type = request.headers.get('Content-Type', '') or ''
        if 'application/json' in content_type:
            try:
                payload = request.body.decode('utf-8') if isinstance(request.body, (bytes, bytearray)) else (request.body or '')
                data = json.loads(payload or '{}')
            except Exception:
                data = {}
            action = data.get('action')
            if action == 'save_selection':
                page_key = data.get('page_key', 'rule_list')
                selected_ids = data.get('selected_ids', [])
                UserTableSelection.objects.update_or_create(
                    user=request.user,
                    page_key=page_key,
                    defaults={'selected_ids': selected_ids}
                )
                return JsonResponse({'success': True, 'count': len(selected_ids)})
            elif action == 'load_selection':
                page_key = data.get('page_key', 'rule_list')
                try:
                    sel = UserTableSelection.objects.get(user=request.user, page_key=page_key)
                    return JsonResponse({'success': True, 'selected_ids': sel.selected_ids})
                except UserTableSelection.DoesNotExist:
                    return JsonResponse({'success': True, 'selected_ids': []})

    # ---- EXPORT SELECTED (POST) ----
    if request.method == 'POST' and 'action' in request.POST:
        action = request.POST.get('action')

        if action in {'export_csv', 'export_excel', 'export_pdf'}:
            _require('export')
        elif action == 'bulk_delete':
            _require('bulk_delete')
        elif action == 'delete_single':
            _require('delete')

        selected_ids = request.POST.getlist('selected_ids')
        if not selected_ids:
            # fallback to persisted selections
            try:
                selection = UserTableSelection.objects.get(user=request.user, page_key='rule_list')
                selected_ids = selection.selected_ids or []
            except UserTableSelection.DoesNotExist:
                selected_ids = []

        if not selected_ids:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'error': 'Tidak ada item yang dipilih'}, status=400)
            messages.error(request, 'Tidak ada item yang dipilih')
            return redirect('manajemen_aplikasi:rule_list')

        qs = PermissionRule.objects.select_related('module','control','function').filter(id__in=selected_ids).annotate(role_count=Count('role_assignments')).order_by('module__order','control__nama_kontrol','function__nama_fungsi')

        if action == 'export_csv':
            return _export_rules_csv(qs)
        elif action == 'export_excel':
            try:
                return _export_rules_excel(qs)
            except Exception:
                return _export_rules_csv(qs)
        elif action == 'export_pdf':
            try:
                return _export_rules_pdf(qs)
            except Exception:
                return _export_rules_csv(qs)
        elif action in {'bulk_delete', 'delete_single'}:
            in_use_ids = set(
                RoleRule.objects.filter(rule_id__in=selected_ids)
                .values_list('rule_id', flat=True)
                .distinct()
            )
            deletable_qs = qs.exclude(id__in=in_use_ids)
            blocked_qs = qs.filter(id__in=in_use_ids)

            deleted_count = deletable_qs.count()
            deleted_names = [str(r) for r in deletable_qs[:50]]
            if deleted_count > 0:
                deletable_qs.delete()
                UserTableSelection.objects.filter(user=request.user, page_key='rule_list').delete()
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({'success': True, 'deleted': True, 'count': deleted_count, 'names': deleted_names})
                messages.success(request, f'{deleted_count} rule berhasil dihapus.')

            blocked_count = blocked_qs.count()
            if blocked_count > 0:
                blocked_names = [str(r) for r in blocked_qs[:50]]
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({'success': False, 'deleted': False, 'blocked': True, 'count': blocked_count, 'names': blocked_names}, status=400)
                messages.warning(request, f'{blocked_count} rule tidak dihapus karena masih terhubung ke Roles.')

            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': True, 'deleted': False, 'count': 0})
            return redirect('manajemen_aplikasi:rule_list')

    rules_qs = PermissionRule.objects.select_related('module', 'control', 'function').annotate(
        role_count=Count('role_assignments'),
        perm_str=Concat('module__nama_module', Value('.'), 'control__nama_kontrol', Value('.'), 'function__nama_fungsi')
    )

    # Optional basic search (by permission string) and module filter
    search_query = (request.GET.get('search') or '').strip()
    if search_query:
        # AND each word across OR fields for more precise results
        terms = [t for t in search_query.split() if t]
        for t in terms:
            rules_qs = rules_qs.filter(
                Q(module__label_module__icontains=t) |
                Q(module__nama_module__icontains=t) |
                Q(control__label_kontrol__icontains=t) |
                Q(control__nama_kontrol__icontains=t) |
                Q(function__label_fungsi__icontains=t) |
                Q(function__nama_fungsi__icontains=t) |
                Q(perm_str__icontains=t)
            )
    module_filter = (request.GET.get('module') or '').strip()
    if module_filter:
        rules_qs = rules_qs.filter(module__nama_module=module_filter)

    rules_qs = rules_qs.order_by('module__order', 'control__nama_kontrol', 'function__nama_fungsi')

    total = rules_qs.count()
    table = RuleTable(rules_qs)
    per_page = request.GET.get('per_page', '10')
    try:
        per_page = int(per_page)
        if per_page not in [10, 25, 50, 100]:
            per_page = 10
    except (ValueError, TypeError):
        per_page = 10
    RequestConfig(request, paginate={'per_page': per_page}).configure(table)

    modules = PermissionModule.objects.all().order_by('order')
    context = {
        'table': table,
        'total': total,
        'modules': modules,
        'current_module': module_filter,
        'search_query': search_query,
    }
    is_htmx = request.headers.get('HX-Request') or request.headers.get('X-Requested-With') == 'XMLHttpRequest'
    if is_htmx:
        return render(request, 'manajemen_aplikasi/access/ma_ap_rules/partials/_table.html', context)
    return render(request, 'manajemen_aplikasi/access/ma_ap_rules/list.html', context)


# ======== EXPORT HELPERS ========
def _export_rules_csv(rules_qs):
    from django.http import HttpResponse
    import csv
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="rules_export.csv"'
    writer = csv.writer(response)
    writer.writerow(['No','ID','Module','Control','Function','Permission String','Status','Roles'])
    for i, r in enumerate(rules_qs, 1):
        writer.writerow([
            i,
            r.id,
            getattr(r.module,'label_module',''),
            getattr(r.control,'label_kontrol',''),
            getattr(r.function,'label_fungsi',''),
            getattr(r, 'permission_string', ''),
            'Active' if r.is_active else 'Inactive',
            getattr(r, 'role_count', 0)
        ])
    return response


def _export_rules_excel(rules_qs):
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    wb = Workbook()
    ws = wb.active
    ws.title = 'Rules Export'
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    headers = ['No','ID','Module','Control','Function','Permission String','Status','Roles']
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    for idx, r in enumerate(rules_qs, start=2):
        data = [idx-1, r.id, getattr(r.module,'label_module',''), getattr(r.control,'label_kontrol',''), getattr(r.function,'label_fungsi',''), getattr(r,'permission_string',''), 'Active' if r.is_active else 'Inactive', getattr(r,'role_count',0)]
        for c, v in enumerate(data, 1):
            cell = ws.cell(row=idx, column=c, value=v)
            cell.border = border
            if c in (1,2,8):
                cell.alignment = Alignment(horizontal='center')
    widths = [6,8,22,24,18,40,10,8]
    from openpyxl.utils import get_column_letter
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return HttpResponse(out.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers={'Content-Disposition':'attachment; filename="rules_export.xlsx"'})


def _export_rules_pdf(rules_qs):
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER
        from io import BytesIO
        buf = BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=18, rightMargin=18, topMargin=24, bottomMargin=24)
        styles = getSampleStyleSheet()
        title = Paragraph('<b>Rules Export</b>', ParagraphStyle('t', parent=styles['Heading1'], alignment=TA_CENTER, fontSize=16))
        data = [['No','ID','Module','Control','Function','Permission','Status','Roles']]
        cell = ParagraphStyle('c', parent=styles['Normal'], fontSize=8, leading=10)
        for i, r in enumerate(rules_qs, 1):
            data.append([str(i), str(r.id), Paragraph(getattr(r.module,'label_module',''), cell), Paragraph(getattr(r.control,'label_kontrol',''), cell), Paragraph(getattr(r.function,'label_fungsi',''), cell), Paragraph(getattr(r,'permission_string',''), cell), 'Active' if r.is_active else 'Inactive', str(getattr(r,'role_count',0))])
        t = Table(data, repeatRows=1)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#666666')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('ALIGN', (0,0), (-1,0), 'CENTER'),
            ('FONTSIZE', (0,0), (-1,0), 10),
            ('TOPPADDING',(0,0),(-1,0),6),
            ('BOTTOMPADDING',(0,0),(-1,0),6),
            ('GRID',(0,0),(-1,-1),0.5, colors.grey),
            ('ROWBACKGROUNDS',(0,1),(-1,-1), [colors.white, colors.HexColor('#f5f5f5')])
        ]))
        elems = [title, Spacer(1,10), t]
        doc.build(elems)
        buf.seek(0)
        return HttpResponse(buf.read(), content_type='application/pdf', headers={'Content-Disposition':'attachment; filename="rules_export.pdf"'})
    except Exception:
        return _export_rules_csv(rules_qs)


@permission_required_403('pengaturan', 'permission_rule', 'create')
def rule_create(request):
    """Create new rule"""
    if request.method == 'POST':
        module_id = request.POST.get('module')
        control_id = request.POST.get('control')
        function_id = request.POST.get('function')
        is_active = request.POST.get('is_active') == 'on'

        errors = []
        field_errors = {}
        module = control = function = None

        # Try to fetch individually for persistence in the form
        selected_module = None
        selected_control = None
        selected_function = None
        if module_id:
            selected_module = PermissionModule.objects.filter(id=module_id).first()
            if not selected_module:
                field_errors.setdefault('module', []).append('Module tidak valid!')
                errors.append('Module tidak valid!')
        else:
            field_errors.setdefault('module', []).append('Module harus dipilih!')
            errors.append('Module harus dipilih!')
        if control_id:
            selected_control = PermissionControl.objects.filter(id=control_id).first()
            if not selected_control:
                field_errors.setdefault('control', []).append('Control tidak valid!')
                errors.append('Control tidak valid!')
        else:
            field_errors.setdefault('control', []).append('Control harus dipilih!')
            errors.append('Control harus dipilih!')
        if function_id:
            selected_function = PermissionFunction.objects.filter(id=function_id).first()
            if not selected_function:
                field_errors.setdefault('function', []).append('Function tidak valid!')
                errors.append('Function tidak valid!')
        else:
            field_errors.setdefault('function', []).append('Function harus dipilih!')
            errors.append('Function harus dipilih!')

        # If all valid selections exist, set them for creation and check duplication
        if selected_module and selected_control and selected_function:
            module = selected_module
            control = selected_control
            function = selected_function

        if module and control and function and PermissionRule.objects.filter(module=module, control=control, function=function).exists():
            errors.append('Rule ini sudah ada!')
            field_errors.setdefault('function', []).append('Rule ini sudah ada!')

        is_htmx = request.headers.get('HX-Request') or request.headers.get('X-Requested-With') == 'XMLHttpRequest'

        modules = PermissionModule.objects.all().order_by('order')
        controls = PermissionControl.objects.all().order_by('nama_kontrol')
        functions = PermissionFunction.objects.all().order_by('nama_fungsi')

        if errors:
            if is_htmx:
                # Do NOT pass errors_list to avoid duplicate summary (JS will render it)
                resp = render(request, 'manajemen_aplikasi/access/ma_ap_rules/form.html', {
                    'modules': modules,
                    'controls': controls,
                    'functions': functions,
                    'selected_module': selected_module,
                    'selected_control': selected_control,
                    'selected_function': selected_function,
                    'is_active_value': is_active,
                })
                resp['HX-Trigger-After-Swap'] = json.dumps({'rule-form-invalid': {'errors': errors, 'fieldErrors': field_errors}})
                return resp
            else:
                for e in errors:
                    messages.error(request, e)
                # Do not pass errors_list here; Django messages will show once
                return render(request, 'manajemen_aplikasi/access/ma_ap_rules/form.html', {
                    'modules': modules,
                    'controls': controls,
                    'functions': functions,
                    'selected_module': selected_module,
                    'selected_control': selected_control,
                    'selected_function': selected_function,
                    'is_active_value': is_active,
                })

        rule = PermissionRule.objects.create(
            module=module,
            control=control,
            function=function,
            is_active=is_active
        )

        if is_htmx:
            resp = HttpResponse(status=204)
            resp['HX-Trigger'] = json.dumps({
                'rule-form-success': {
                    'message': f'Rule "{rule}" berhasil dibuat!',
                    'redirect': reverse('manajemen_aplikasi:rule_list')
                }
            })
            return resp

        messages.success(request, f'✅ Rule "{rule}" berhasil dibuat!')
        return redirect('manajemen_aplikasi:rule_list')

    # Get all data for form
    modules = PermissionModule.objects.all().order_by('order')
    controls = PermissionControl.objects.all().order_by('nama_kontrol')
    functions = PermissionFunction.objects.all().order_by('nama_fungsi')

    context = {
        'modules': modules,
        'controls': controls,
        'functions': functions,
        'is_active_value': True,
    }
    return render(request, 'manajemen_aplikasi/access/ma_ap_rules/form.html', context)


@permission_required_403('pengaturan', 'permission_rule', 'edit')
def rule_edit(request, rule_id):
    """Edit rule"""
    rule = get_object_or_404(PermissionRule, id=rule_id)

    if request.method == 'POST':
        module_id = request.POST.get('module')
        control_id = request.POST.get('control')
        function_id = request.POST.get('function')
        is_active = request.POST.get('is_active') == 'on'

        errors = []
        field_errors = {}

        selected_module = None
        selected_control = None
        selected_function = None
        if module_id:
            selected_module = PermissionModule.objects.filter(id=module_id).first()
            if not selected_module:
                field_errors.setdefault('module', []).append('Module tidak valid!')
                errors.append('Module tidak valid!')
        else:
            field_errors.setdefault('module', []).append('Module harus dipilih!')
            errors.append('Module harus dipilih!')
        if control_id:
            selected_control = PermissionControl.objects.filter(id=control_id).first()
            if not selected_control:
                field_errors.setdefault('control', []).append('Control tidak valid!')
                errors.append('Control tidak valid!')
        else:
            field_errors.setdefault('control', []).append('Control harus dipilih!')
            errors.append('Control harus dipilih!')
        if function_id:
            selected_function = PermissionFunction.objects.filter(id=function_id).first()
            if not selected_function:
                field_errors.setdefault('function', []).append('Function tidak valid!')
                errors.append('Function tidak valid!')
        else:
            field_errors.setdefault('function', []).append('Function harus dipilih!')
            errors.append('Function harus dipilih!')

        # Duplicate check excluding current rule
        if selected_module and selected_control and selected_function:
            exists = PermissionRule.objects.filter(module=selected_module, control=selected_control, function=selected_function).exclude(pk=rule.id).exists()
            if exists:
                dup_msg = 'Rule ini sudah ada!'
                errors.append(dup_msg)
                field_errors.setdefault('function', []).append(dup_msg)

        is_htmx = request.headers.get('HX-Request') or request.headers.get('X-Requested-With') == 'XMLHttpRequest'

        modules = PermissionModule.objects.all().order_by('order')

        if errors:
            if is_htmx:
                ctx = {
                    'rule': rule,
                    'edit_mode': True,
                    'modules': modules,
                    'selected_module': selected_module,
                    'selected_control': selected_control,
                    'selected_function': selected_function,
                }
                resp = render(request, 'manajemen_aplikasi/access/ma_ap_rules/form.html', ctx)
                resp['HX-Trigger-After-Swap'] = json.dumps({'rule-form-invalid': {'errors': errors, 'fieldErrors': field_errors}})
                return resp
            else:
                for e in errors:
                    messages.error(request, e)
                return render(request, 'manajemen_aplikasi/access/ma_ap_rules/form.html', {
                    'rule': rule,
                    'edit_mode': True,
                    'modules': modules,
                    'selected_module': selected_module,
                    'selected_control': selected_control,
                    'selected_function': selected_function,
                })

        # Apply updates
        try:
            if selected_module: rule.module = selected_module
            if selected_control: rule.control = selected_control
            if selected_function: rule.function = selected_function
            rule.is_active = is_active
            rule.save()
        except Exception:
            messages.error(request, 'Gagal menyimpan perubahan.')
            return redirect('manajemen_aplikasi:rule_edit', rule_id=rule.id)

        if is_htmx:
            resp = HttpResponse(status=204)
            resp['HX-Trigger'] = json.dumps({'rule-form-success': {'message': f'Rule "{rule}" berhasil diupdate!', 'redirect': reverse('manajemen_aplikasi:rule_list')}})
            return resp
        messages.success(request, f'✅ Rule "{rule}" berhasil diupdate!')
        return redirect('manajemen_aplikasi:rule_list')

    modules = PermissionModule.objects.all().order_by('order')
    context = {
        'rule': rule,
        'edit_mode': True,
        'modules': modules,
    }
    return render(request, 'manajemen_aplikasi/access/ma_ap_rules/form.html', context)


@permission_required_403('pengaturan', 'permission_rule', 'delete')
def rule_delete(request, rule_id):
    """Delete rule"""
    rule = get_object_or_404(PermissionRule, id=rule_id)

    if request.method == 'POST':
        rule_str = str(rule)
        rule.delete()
        # AJAX friendly: return JSON if request made via XHR
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': True, 'deleted': rule_id, 'name': rule_str})
        messages.success(request, f'✅ Rule "{rule_str}" berhasil dihapus!')
        return redirect('manajemen_aplikasi:rule_list')

    # Get affected role assignments
    affected_assignments = RoleRule.objects.filter(rule=rule).select_related('role')

    context = {
        'rule': rule,
        'affected_assignments': affected_assignments,
        'assignment_count': affected_assignments.count(),
    }
    return render(request, 'manajemen_aplikasi/access/ma_ap_rules/delete.html', context)


# =============================
# AJAX SEARCH ENDPOINTS (Remote)
# =============================
@permission_required_403('pengaturan', 'permission_rule', 'view')
def ajax_controls_search(request):
    """Return paginated controls for remote searchable select.
    Query params: q, page (1-based), page_size.
    Response: {results: [{value, label}], has_more: bool}
    """
    q = (request.GET.get('q') or '').strip()
    try:
        page = max(1, int(request.GET.get('page') or 1))
    except Exception:
        page = 1
    try:
        page_size = max(1, min(200, int(request.GET.get('page_size') or 50)))
    except Exception:
        page_size = 50

    qs = PermissionControl.objects.all()
    if q:
        qs = qs.filter(Q(nama_kontrol__icontains=q) | Q(label_kontrol__icontains=q))
    qs = qs.order_by('label_kontrol', 'nama_kontrol')

    start = (page - 1) * page_size
    # Fetch one extra to detect has_more without COUNT(*).
    rows = list(qs[start:start + page_size + 1])
    has_more = len(rows) > page_size
    rows = rows[:page_size]

    data = [{
        'value': r.id,
        'label': f"{r.label_kontrol} ({r.nama_kontrol})"
    } for r in rows]
    return JsonResponse({'results': data, 'has_more': has_more})


@permission_required_403('pengaturan', 'permission_rule', 'view')
def ajax_functions_search(request):
    """Return paginated functions for remote searchable select.
    Query params: q, page (1-based), page_size.
    Response: {results: [{value, label}], has_more: bool}
    """
    q = (request.GET.get('q') or '').strip()
    try:
        page = max(1, int(request.GET.get('page') or 1))
    except Exception:
        page = 1
    try:
        page_size = max(1, min(200, int(request.GET.get('page_size') or 50)))
    except Exception:
        page_size = 50

    qs = PermissionFunction.objects.all()
    if q:
        qs = qs.filter(Q(nama_fungsi__icontains=q) | Q(label_fungsi__icontains=q))
    qs = qs.order_by('label_fungsi', 'nama_fungsi')

    start = (page - 1) * page_size
    rows = list(qs[start:start + page_size + 1])
    has_more = len(rows) > page_size
    rows = rows[:page_size]

    data = [{
        'value': r.id,
        'label': f"{r.label_fungsi} ({r.nama_fungsi})"
    } for r in rows]
    return JsonResponse({'results': data, 'has_more': has_more})
