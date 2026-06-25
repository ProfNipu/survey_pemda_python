from django.shortcuts import render, redirect, get_object_or_404
from django.views.decorators.csrf import ensure_csrf_cookie
from django.contrib import messages
from django.db.models import Q, Count
from django_tables2 import RequestConfig
from django.http import HttpResponse
from django.urls import reverse
import json

from apps.manajemen.decorators import permission_required_403
from apps.manajemen.models import MenuCategory, MenuItem
from apps.manajemen.forms import MenuCategoryForm
from apps.manajemen.tables import MenuCategoryTable


def export_menu_categories_csv(categories):
    """Export menu categories to CSV"""
    import csv
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="menu_categories_export.csv"'
    writer = csv.writer(response)
    writer.writerow(['Code', 'Nama Kategori', 'Order', 'Active', 'Jumlah Menu'])
    for c in categories:
        writer.writerow([
            c.code,
            c.name,
            c.order,
            'Yes' if c.is_active else 'No',
            getattr(c, 'menus_count', 0),
        ])
    return response


def export_menu_categories_excel(categories):
    """Export menu categories to Excel (openpyxl if available, else CSV)"""
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
        from io import BytesIO
        wb = Workbook()
        ws = wb.active
        ws.title = 'Menu Categories'
        headers = ['Code', 'Nama Kategori', 'Order', 'Active', 'Jumlah Menu']
        ws.append(headers)
        for c in categories:
            ws.append([
                c.code,
                c.name,
                c.order,
                'Yes' if c.is_active else 'No',
                getattr(c, 'menus_count', 0),
            ])
        for col_idx, _ in enumerate(headers, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 22
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="menu_categories_export.xlsx"'
        return response
    except ImportError:
        return export_menu_categories_csv(categories)


def export_menu_categories_pdf(categories):
    """Export menu categories to PDF (A4 portrait). Falls back to CSV if reportlab not available"""
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas
        from reportlab.lib.units import cm
        from io import BytesIO

        buffer = BytesIO()
        p = canvas.Canvas(buffer, pagesize=A4)
        width, height = A4

        y = height - 2 * cm
        p.setFont('Helvetica-Bold', 14)
        p.drawString(2 * cm, y, 'Menu Categories Export')
        y -= 1 * cm

        p.setFont('Helvetica', 10)
        headers = ['Code', 'Nama Kategori', 'Order', 'Active', 'Jumlah Menu']
        col_x = [2 * cm, 5 * cm, 13 * cm, 16 * cm, 19 * cm]
        # Header
        for i, h in enumerate(headers):
            p.drawString(col_x[i], y, h)
        y -= 0.6 * cm

        # Rows
        for c in categories:
            if y < 2 * cm:
                p.showPage()
                y = height - 2 * cm
                p.setFont('Helvetica', 10)
                for i, h in enumerate(headers):
                    p.drawString(col_x[i], y, h)
                y -= 0.6 * cm
            row = [
                str(c.code),
                str(c.name),
                str(c.order),
                'Yes' if c.is_active else 'No',
                str(getattr(c, 'menus_count', 0)),
            ]
            for i, val in enumerate(row):
                p.drawString(col_x[i], y, val)
            y -= 0.5 * cm

        p.showPage()
        p.save()
        pdf = buffer.getvalue()
        buffer.close()
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="menu_categories_export.pdf"'
        response.write(pdf)
        return response
    except ImportError:
        return export_menu_categories_csv(categories)


@ensure_csrf_cookie
@permission_required_403('pengaturan', 'menu_category', 'view')
def menu_category_list(request):
    qs = MenuCategory.objects.all()
    search_query = request.GET.get('search', '').strip()
    # Handle EXPORT ALL on POST
    if request.method == 'POST' and 'export_all' in request.POST:
        export_format = request.POST.get('export_all')
        search_post = request.POST.get('search', '').strip()
        exp_qs = MenuCategory.objects.all()
        if search_post:
            exp_qs = exp_qs.filter(Q(name__icontains=search_post) | Q(code__icontains=search_post))
        exp_qs = exp_qs.order_by('order', 'name')

        # Attach menus_count
        counts_qs = MenuItem.objects.values('category').annotate(c=Count('id'))
        counts_map = {row['category']: row['c'] for row in counts_qs}
        exp_list = list(exp_qs)
        for obj in exp_list:
            setattr(obj, 'menus_count', counts_map.get(obj.code, 0))

        if export_format == 'excel':
            try:
                return export_menu_categories_excel(exp_list)
            except Exception:
                return export_menu_categories_csv(exp_list)
        elif export_format == 'pdf':
            try:
                return export_menu_categories_pdf(exp_list)
            except Exception:
                return export_menu_categories_csv(exp_list)
        else:
            return export_menu_categories_csv(exp_list)
    if search_query:
        qs = qs.filter(
            Q(name__icontains=search_query) |
            Q(code__icontains=search_query)
        )
    qs = qs.order_by('order', 'name')

    total = qs.count()

    counts_qs = (
        MenuItem.objects.values('category')
        .annotate(c=Count('id'))
    )
    counts_map = {row['category']: row['c'] for row in counts_qs}

    qs_list = list(qs)
    for obj in qs_list:
        setattr(obj, 'menus_count', counts_map.get(obj.code, 0))

    table = MenuCategoryTable(qs_list)
    per_page = request.GET.get('per_page', '10')
    try:
        per_page = int(per_page)
        if per_page not in [10, 25, 50, 100]:
            per_page = 10
    except (ValueError, TypeError):
        per_page = 10
    RequestConfig(request, paginate={'per_page': per_page}).configure(table)

    context = {
        'table': table,
        'total': total,
        'search_query': search_query,
    }
    is_htmx = request.headers.get('HX-Request') or request.headers.get('X-Requested-With') == 'XMLHttpRequest'
    if is_htmx:
        return render(request, 'manajemen_aplikasi/access/ma_ap_menu_categories/partials/_table.html', context)
    return render(request, 'manajemen_aplikasi/access/ma_ap_menu_categories/list.html', context)


@permission_required_403('pengaturan', 'menu_category', 'create')
def menu_category_create(request):
    if request.method == 'POST':
        form = MenuCategoryForm(request.POST)
        is_htmx = request.headers.get('HX-Request') or request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        if not form.is_valid():
            errors = []
            field_errors = {}
            for field, field_errors in form.errors.items():
                errs_list = []
                for e in field_errors:
                    s = str(e)
                    lbl = {'code': 'Kode', 'name': 'Nama', 'order': 'Order'}.get(field, field)
                    sl = s.lower()
                    if sl in ('bidang ini harus diisi.', 'this field is required.'):
                        s = f'{lbl} harus diisi.'
                    if ('already exists' in sl or 'telah ada' in sl) and field == 'code':
                        s = 'Kode sudah digunakan.'
                    errors.append(s)
                    errs_list.append(s)
                try:
                    field_errors[field] = errs_list
                except Exception:
                    pass
            non_field = form.non_field_errors()
            if non_field:
                errors.extend([str(e) for e in non_field])

            if is_htmx:
                resp = render(request, 'manajemen_aplikasi/access/ma_ap_menu_categories/form.html', {'form': form, 'errors_list': errors})
                resp['HX-Trigger-After-Swap'] = json.dumps({'menu-category-form-invalid': {'errors': errors, 'fieldErrors': field_errors}})
                return resp
            else:
                for e in errors:
                    messages.error(request, e)
                return render(request, 'manajemen_aplikasi/access/ma_ap_menu_categories/form.html', {'form': form, 'errors_list': errors})

        obj = form.save()
        if is_htmx:
            resp = HttpResponse(status=204)
            resp['HX-Trigger'] = json.dumps({'menu-category-form-success': {'message': f'Kategori "{obj.name}" berhasil dibuat', 'redirect': reverse('manajemen_aplikasi:menu_category_list')}})
            return resp
        messages.success(request, 'Kategori berhasil dibuat')
        return redirect('manajemen_aplikasi:menu_category_list')
    else:
        form = MenuCategoryForm()
    return render(request, 'manajemen_aplikasi/access/ma_ap_menu_categories/form.html', {'form': form})


@permission_required_403('pengaturan', 'menu_category', 'edit')
def menu_category_edit(request, cat_id):
    obj = get_object_or_404(MenuCategory, id=cat_id)
    if request.method == 'POST':
        form = MenuCategoryForm(request.POST, instance=obj)
        is_htmx = request.headers.get('HX-Request') or request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        if not form.is_valid():
            errors = []
            field_errors = {}
            for field, field_errors in form.errors.items():
                errs_list = []
                for e in field_errors:
                    s = str(e)
                    lbl = {'code': 'Kode', 'name': 'Nama', 'order': 'Order'}.get(field, field)
                    sl = s.lower()
                    if sl in ('bidang ini harus diisi.', 'this field is required.'):
                        s = f'{lbl} harus diisi.'
                    if ('already exists' in sl or 'telah ada' in sl) and field == 'code':
                        s = 'Kode sudah digunakan.'
                    errors.append(s)
                    errs_list.append(s)
                try:
                    field_errors[field] = errs_list
                except Exception:
                    pass
            non_field = form.non_field_errors()
            if non_field:
                errors.extend([str(e) for e in non_field])

            if is_htmx:
                resp = render(request, 'manajemen_aplikasi/access/ma_ap_menu_categories/form.html', {'form': form, 'edit_mode': True, 'category': obj, 'errors_list': errors})
                resp['HX-Trigger-After-Swap'] = json.dumps({'menu-category-form-invalid': {'errors': errors, 'fieldErrors': field_errors}})
                return resp
            else:
                for e in errors:
                    messages.error(request, e)
                return render(request, 'manajemen_aplikasi/access/ma_ap_menu_categories/form.html', {'form': form, 'edit_mode': True, 'category': obj, 'errors_list': errors})

        obj = form.save()
        if is_htmx:
            resp = HttpResponse(status=204)
            resp['HX-Trigger'] = json.dumps({'menu-category-form-success': {'message': f'Kategori "{obj.name}" berhasil diupdate', 'redirect': reverse('manajemen_aplikasi:menu_category_list')}})
            return resp
        messages.success(request, 'Kategori berhasil diupdate')
        return redirect('manajemen_aplikasi:menu_category_list')
    else:
        form = MenuCategoryForm(instance=obj)
    return render(request, 'manajemen_aplikasi/access/ma_ap_menu_categories/form.html', {'form': form, 'edit_mode': True, 'category': obj})


@permission_required_403('pengaturan', 'menu_category', 'delete')
def menu_category_delete(request, cat_id):
    obj = get_object_or_404(MenuCategory, id=cat_id)
    if request.method == 'POST':
        linked = MenuItem.objects.filter(category=obj.code).count()
        if linked > 0:
            messages.error(request, f'Kategori "{obj.name}" tidak dapat dihapus karena dipakai oleh {linked} menu')
            return redirect('manajemen_aplikasi:menu_category_list')
        name = obj.name
        obj.delete()
        messages.success(request, f'Kategori "{name}" berhasil dihapus')
        return redirect('manajemen_aplikasi:menu_category_list')
    linked = MenuItem.objects.filter(category=obj.code).count()
    if linked > 0:
        messages.error(request, f'Kategori "{obj.name}" tidak dapat dihapus karena dipakai oleh {linked} menu')
        return redirect('manajemen_aplikasi:menu_category_list')
    return render(request, 'manajemen_aplikasi/access/ma_ap_menu_categories/delete.html', {'category': obj})
