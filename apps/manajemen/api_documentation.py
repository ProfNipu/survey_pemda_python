from __future__ import annotations

import json
import re
from pathlib import Path

from django.contrib import messages
from django.conf import settings
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.views.decorators.csrf import ensure_csrf_cookie
from django.db.models import Q
from django_tables2 import RequestConfig
import csv
from io import BytesIO

from apps.manajemen.decorators import permission_required_403
from apps.manajemen.models import ApiDocumentation, UserTableSelection
from apps.manajemen.tables import ApiDocumentationTable


def _normalize_method(value: str | None) -> str:
    if not value:
        return ''
    s = str(value).strip()
    if not s:
        return ''
    try:
        loaded = json.loads(s)
        if isinstance(loaded, list):
            return ', '.join([str(x) for x in loaded])
        return str(loaded)
    except Exception:
        return s


def _get_siasn_api_docs_seed() -> list[dict]:
    return [
        {
            'method_type': 'GET',
            'url': '/integrations/pegawai/fetch/<nip>/',
            'description': 'Fetch & cache data SIASN (default: integrasi/full) untuk NIP tertentu. Menyimpan hasil ke tabel siasn_pegawai + progress tracking.',
            'parameters': {
                'path': {'nip': 'NIP 18 digit'},
                'fetch_type': 'integrasi (default)'}
        },
        {
            'method_type': 'GET',
            'url': '/integrations/pegawai/fetch/<nip>/<fetch_type>/',
            'description': 'Fetch & cache data SIASN untuk NIP tertentu dengan tipe fetch. minimal/sync mengambil subset endpoint untuk lebih cepat.',
            'parameters': {
                'path': {
                    'nip': 'NIP 18 digit',
                    'fetch_type': 'integrasi | minimal | sync'
                }
            }
        },
        {
            'method_type': 'GET',
            'url': '/integrations/pegawai/progress/<nip>/',
            'description': 'Ambil progress fetch SIASN per NIP (real-time progress tracker).',
            'parameters': {'path': {'nip': 'NIP 18 digit'}}
        },
        {
            'method_type': 'GET',
            'url': '/integrations/pegawai/progress/<nip>/<fetch_type>/',
            'description': 'Ambil progress fetch SIASN per NIP dan tipe fetch.',
            'parameters': {'path': {'nip': 'NIP 18 digit', 'fetch_type': 'integrasi | minimal | sync'}}
        },
        {
            'method_type': 'GET',
            'url': '/integrations/pegawai/list/',
            'description': 'List data cache SIASN (tabel siasn_pegawai).',
            'parameters': None
        },
        {
            'method_type': 'GET',
            'url': '/integrations/api-logs/',
            'description': 'Lihat log panggilan API SIASN (SiasnApiLog).',
            'parameters': None
        },
        {
            'method_type': 'POST',
            'url': '/integrations/siasn/token/refresh/',
            'description': 'Refresh token SIASN (SSO + API token) sesuai mekanisme integrasi.',
            'parameters': None
        },
        {
            'method_type': 'GET',
            'url': '/integrations/siasn/download-dokumen/?uri=<path>',
            'description': 'Download dokumen SIASN (base64) berdasarkan URI/path dokumen.',
            'parameters': {'query': {'uri': 'dokumen URI/path dari response SIASN'}}
        },
    ]


def _get_internal_ajax_api_docs_seed() -> list[dict]:
    return [
        {
            'method_type': 'GET',
            'url': '/manajemen-relasi-organisasi/ajax/jenis-organisasi/',
            'description': 'Remote searchable select: Jenis Organisasi (opsional filter only_opd untuk OPD induk).',
            'parameters': {'query': {'q': 'search', 'page': '1-based', 'page_size': 'page size', 'only_opd': '0/1'}}
        },
        {
            'method_type': 'GET',
            'url': '/manajemen-relasi-organisasi/ajax/unit-kerja/',
            'description': 'Remote searchable select: Unit Kerja (OPD induk).',
            'parameters': {'query': {'q': 'search', 'page': '1-based', 'page_size': 'page size'}}
        },
        {
            'method_type': 'GET',
            'url': '/manajemen-relasi-organisasi/ajax/sub-unit-kerja/',
            'description': 'Remote searchable select: Sub Unit Kerja (tree) berdasarkan parent_id.',
            'parameters': {'query': {'parent_id': 'required (numeric)', 'q': 'search', 'page': '1-based', 'page_size': 'page size'}}
        },
        {
            'method_type': 'GET',
            'url': '/manajemen-relasi-organisasi/ajax/jabatan-struktural/',
            'description': 'Remote searchable select: Jabatan Struktural (filter by id_opd / id_sub_opd).',
            'parameters': {'query': {'id_opd': 'required (numeric)', 'id_sub_opd': 'optional (numeric)', 'q': 'search', 'page': '1-based', 'page_size': 'page size'}}
        },
        {
            'method_type': 'GET',
            'url': '/manajemen-relasi-organisasi/ajax/jabatan-non-struktural/',
            'description': 'Remote searchable select: Jabatan Non Struktural (filter by id_jenis_jabatan).',
            'parameters': {'query': {'id_jenis_jabatan': 'required (numeric)', 'q': 'search', 'page': '1-based', 'page_size': 'page size'}}
        },
        {
            'method_type': 'GET',
            'url': '/manajemen-relasi-organisasi/ajax/bkn-jabatan-fungsional/',
            'description': 'Remote searchable select: BKN Jabatan Fungsional (untuk Jenis Jabatan=3).',
            'parameters': {'query': {'q': 'search', 'page': '1-based', 'page_size': 'page size'}}
        },
        {
            'method_type': 'GET',
            'url': '/manajemen-relasi-organisasi/ajax/bkn-sub-jabatan/',
            'description': 'Remote searchable select: BKN Sub Jabatan (berdasarkan id_jabatan[] yang dipilih).',
            'parameters': {'query': {'id_jabatan': 'list (repeat param) required', 'q': 'search', 'page': '1-based', 'page_size': 'page size'}}
        },
        {
            'method_type': 'GET',
            'url': '/manajemen-aplikasi/rules/ajax/controls/',
            'description': 'Remote searchable select: list Permission Controls untuk Rules builder.',
            'parameters': {'query': {'q': 'search', 'page': '1-based', 'page_size': 'page size'}}
        },
        {
            'method_type': 'GET',
            'url': '/manajemen-aplikasi/rules/ajax/functions/',
            'description': 'Remote searchable select: list Permission Functions untuk Rules builder.',
            'parameters': {'query': {'q': 'search', 'page': '1-based', 'page_size': 'page size'}}
        },
    ]


def export_api_documentation_csv(qs):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="api_documentation_export.csv"'
    writer = csv.writer(response)
    writer.writerow(['No', 'Method', 'URL', 'Description', 'Active'])
    for i, obj in enumerate(qs, 1):
        writer.writerow([
            i,
            obj.method_type,
            obj.url,
            obj.description or '',
            'Yes' if obj.is_active else 'No',
        ])
    return response


def export_api_documentation_excel(qs):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = Workbook()
        ws = wb.active
        ws.title = 'API Documentation'

        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        headers = ['No', 'Method', 'URL', 'Description', 'Active']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        for row_idx, obj in enumerate(qs, 2):
            ws.cell(row=row_idx, column=1, value=row_idx - 1)
            ws.cell(row=row_idx, column=2, value=obj.method_type)
            ws.cell(row=row_idx, column=3, value=obj.url)
            ws.cell(row=row_idx, column=4, value=obj.description or '')
            ws.cell(row=row_idx, column=5, value='Yes' if obj.is_active else 'No')

        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 50
        ws.column_dimensions['D'].width = 60
        ws.column_dimensions['E'].width = 10

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        resp = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        resp['Content-Disposition'] = 'attachment; filename="api_documentation_export.xlsx"'
        return resp
    except ImportError:
        return export_api_documentation_csv(qs)


@permission_required_403('pengaturan', 'api_documentation', 'create')
def api_documentation_seed_siasn_endpoints(request):
    created = 0
    skipped = 0
    for item in _get_siasn_api_docs_seed():
        method_type = (item.get('method_type') or '').strip().upper()
        url = (item.get('url') or '').strip()
        if not method_type or not url:
            continue
        if ApiDocumentation.objects.filter(method_type=method_type, url=url).exists():
            skipped += 1
            continue
        ApiDocumentation.objects.create(
            method_type=method_type,
            url=url,
            description=item.get('description') or '',
            parameters=item.get('parameters'),
            is_active=True,
        )
        created += 1

    messages.success(request, f'{created} endpoint SIASN ditambahkan. {skipped} sudah ada (skip).')
    return redirect('manajemen_aplikasi:api_documentation_list')


@permission_required_403('pengaturan', 'api_documentation', 'create')
def api_documentation_seed_internal_ajax_endpoints(request):
    created = 0
    skipped = 0
    updated = 0

    for item in _get_internal_ajax_api_docs_seed():
        method_type = (item.get('method_type') or '').strip().upper()
        url = (item.get('url') or '').strip()
        if not method_type or not url:
            continue

        defaults = {
            'description': item.get('description') or '',
            'parameters': item.get('parameters'),
            'is_active': True,
        }

        obj, was_created = ApiDocumentation.objects.get_or_create(
            method_type=method_type,
            url=url,
            defaults=defaults,
        )
        if was_created:
            created += 1
            continue

        changed = False
        for field, val in defaults.items():
            if getattr(obj, field) != val:
                setattr(obj, field, val)
                changed = True
        if changed:
            obj.save(update_fields=['description', 'parameters', 'is_active', 'updated_at'])
            updated += 1
        else:
            skipped += 1

    messages.success(request, f'{created} endpoint AJAX internal ditambahkan. {updated} diupdate. {skipped} sudah up-to-date (skip).')
    return redirect('manajemen_aplikasi:api_documentation_list')


def _load_api_docs_from_markdown() -> list[dict]:
    md_path = Path(getattr(settings, 'BASE_DIR', Path.cwd())) / 'docs' / 'API-ENDPOINTS.md'
    if not md_path.exists():
        return []

    text = md_path.read_text(encoding='utf-8', errors='ignore')
    lines = text.splitlines()

    docs: list[dict] = []

    current_title: str | None = None
    current_url: str | None = None
    current_method: str | None = None
    current_parameters: list[str] = []
    current_description: str | None = None

    def _flush():
        nonlocal current_title, current_url, current_method, current_parameters, current_description
        if not current_url and not current_method and not current_title:
            return
        if not current_url:
            return
        params_text = ''
        if current_parameters:
            params_text = '\n'.join(current_parameters)
        docs.append(
            {
                'id': len(docs) + 1,
                'title': current_title or '',
                'method_type': current_method or '',
                'url': current_url,
                'parameters': params_text,
                'description': current_description or current_title or '',
                'created_at': None,
                'updated_at': None,
            }
        )
        current_title = None
        current_url = None
        current_method = None
        current_parameters = []
        current_description = None

    in_params = False
    for raw in lines:
        line = raw.strip()

        # New endpoint header
        m = re.match(r'^####\s+\d+\.\s+(.*)$', line)
        if m:
            _flush()
            current_title = m.group(1).strip()
            in_params = False
            continue

        # URL line: **URL:** `/path`
        m = re.match(r'^\*\*URL:\*\*\s+`([^`]+)`\s*$', line)
        if m:
            current_url = m.group(1).strip()
            in_params = False
            continue

        # Method line: **Method:** `GET`
        m = re.match(r'^\*\*Method:\*\*\s+(.+?)\s*$', line)
        if m:
            method_raw = m.group(1).strip()
            # strip backticks, keep things like "GET or POST"
            method_raw = method_raw.replace('`', '')
            current_method = method_raw
            in_params = False
            continue

        if line.startswith('**Parameters:**'):
            in_params = True
            continue

        if in_params:
            if not line:
                in_params = False
                continue
            if line.startswith('- '):
                current_parameters.append(line[2:].strip())
                continue
            # Any other line ends params section
            in_params = False

        # Optional description line (keep minimal). We treat a non-empty paragraph line after title as description
        if current_title and not current_description:
            if line and not line.startswith('**') and not line.startswith('```') and not line.startswith('---'):
                current_description = line

    _flush()

    # Normalize computed fields for templates
    for d in docs:
        d['method_type_display'] = _normalize_method(d.get('method_type'))
        d['parameters_display'] = _normalize_parameters(d.get('parameters'))

    return docs


def _normalize_parameters(value) -> str:
    if value is None:
        return ''
    if isinstance(value, (dict, list)):
        try:
            return json.dumps(value, ensure_ascii=False, indent=2)
        except Exception:
            return str(value)

    s = str(value).strip()
    if not s:
        return ''
    try:
        loaded = json.loads(s)
        return json.dumps(loaded, ensure_ascii=False, indent=2)
    except Exception:
        return s


@ensure_csrf_cookie
@permission_required_403('pengaturan', 'api_documentation', 'view')
def api_documentation_list(request):
    """List API documentation with reusable datatable + selection persistence + bulk actions"""

    # ---- EXPORT ALL (POST) ----
    if request.method == 'POST' and 'export_all' in request.POST:
        export_format = request.POST.get('export_all')
        search_query = request.POST.get('search', '').strip()
        qs = ApiDocumentation.objects.exclude(url__icontains='/ajax/')
        if search_query:
            qs = qs.filter(
                Q(url__icontains=search_query)
                | Q(method_type__icontains=search_query)
                | Q(description__icontains=search_query)
            )
        qs = qs.order_by('-id')
        if export_format == 'excel':
            return export_api_documentation_excel(qs)
        return export_api_documentation_csv(qs)

    # ---- AJAX SAVE/LOAD SELECTIONS ----
    if request.method == 'POST' and request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        content_type = request.headers.get('Content-Type', '') or ''
        if 'application/json' in content_type:
            try:
                payload = request.body.decode('utf-8') if isinstance(request.body, (bytes, bytearray)) else (request.body or '')
                data = json.loads(payload or '{}')
            except Exception:
                data = {}
            action = data.get('action')
            if action == 'save_selection':
                page_key = data.get('page_key', 'api_documentation_list')
                selected_ids = data.get('selected_ids', [])
                UserTableSelection.objects.update_or_create(
                    user=request.user,
                    page_key=page_key,
                    defaults={'selected_ids': selected_ids},
                )
                return JsonResponse({'success': True, 'count': len(selected_ids)})
            if action == 'load_selection':
                page_key = data.get('page_key', 'api_documentation_list')
                try:
                    selection = UserTableSelection.objects.get(user=request.user, page_key=page_key)
                    return JsonResponse({'success': True, 'selected_ids': selection.selected_ids})
                except UserTableSelection.DoesNotExist:
                    return JsonResponse({'success': True, 'selected_ids': []})

    # ---- BULK ACTIONS (POST) ----
    if request.method == 'POST' and 'action' in request.POST:
        action = request.POST.get('action')
        selected_ids = request.POST.getlist('selected_ids')
        if not selected_ids:
            single_id = request.POST.get('id')
            if single_id:
                selected_ids = [single_id]
        if not selected_ids:
            try:
                selection = UserTableSelection.objects.get(user=request.user, page_key='api_documentation_list')
                selected_ids = selection.selected_ids or []
            except UserTableSelection.DoesNotExist:
                selected_ids = []
        if not selected_ids:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'error': 'Tidak ada item yang dipilih'}, status=400)
            messages.error(request, 'Tidak ada item yang dipilih')
            return redirect('manajemen_aplikasi:api_documentation_list')

        qs = ApiDocumentation.objects.filter(id__in=selected_ids).exclude(url__icontains='/ajax/').order_by('-id')
        if action == 'export_csv':
            return export_api_documentation_csv(qs)
        if action == 'export_excel':
            return export_api_documentation_excel(qs)
        if action in ('bulk_delete', 'delete_single'):
            deleted_count = qs.count()
            deleted_urls = list(qs.values_list('url', flat=True))
            if deleted_count:
                qs.delete()
                UserTableSelection.objects.filter(user=request.user, page_key='api_documentation_list').delete()
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': True, 'deleted': True, 'count': deleted_count, 'names': deleted_urls})
            messages.success(request, f'{deleted_count} endpoint berhasil dihapus')
            return redirect('manajemen_aplikasi:api_documentation_list')

    # ---- REGULAR GET ----
    qs = ApiDocumentation.objects.exclude(url__icontains='/ajax/')
    search_query = request.GET.get('search', '').strip()
    if search_query:
        qs = qs.filter(
            Q(url__icontains=search_query)
            | Q(method_type__icontains=search_query)
            | Q(description__icontains=search_query)
        )
    qs = qs.order_by('-id')

    total = qs.count()
    table = ApiDocumentationTable(qs)
    per_page = request.GET.get('per_page', '10')
    try:
        per_page = int(per_page)
        if per_page not in [10, 25, 50, 100]:
            per_page = 10
    except (ValueError, TypeError):
        per_page = 10
    RequestConfig(request, paginate={'per_page': per_page}).configure(table)

    context = {
        'table': table,
        'total': total,
        'search_query': search_query,
        'per_page': per_page,
    }

    is_htmx = request.headers.get('HX-Request') or request.headers.get('X-Requested-With') == 'XMLHttpRequest'
    if is_htmx:
        return render(request, 'manajemen_aplikasi/access/ma_ap_api_documentation/partials/_table.html', context)
    return render(request, 'manajemen_aplikasi/access/ma_ap_api_documentation/list.html', context)


@ensure_csrf_cookie
@permission_required_403('pengaturan', 'api_documentation', 'view')
def ajax_documentation_list(request):
    """List AJAX (internal) endpoints documentation only (URL contains /ajax/)."""

    if request.method == 'POST' and 'export_all' in request.POST:
        export_format = request.POST.get('export_all')
        search_query = request.POST.get('search', '').strip()
        qs = ApiDocumentation.objects.filter(url__icontains='/ajax/')
        if search_query:
            qs = qs.filter(
                Q(url__icontains=search_query)
                | Q(method_type__icontains=search_query)
                | Q(description__icontains=search_query)
            )
        qs = qs.order_by('-id')
        if export_format == 'excel':
            return export_api_documentation_excel(qs)
        return export_api_documentation_csv(qs)

    if request.method == 'POST' and request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        content_type = request.headers.get('Content-Type', '') or ''
        if 'application/json' in content_type:
            try:
                payload = request.body.decode('utf-8') if isinstance(request.body, (bytes, bytearray)) else (request.body or '')
                data = json.loads(payload or '{}')
            except Exception:
                data = {}
            action = data.get('action')
            if action == 'save_selection':
                page_key = data.get('page_key', 'ajax_documentation_list')
                selected_ids = data.get('selected_ids', [])
                UserTableSelection.objects.update_or_create(
                    user=request.user,
                    page_key=page_key,
                    defaults={'selected_ids': selected_ids},
                )
                return JsonResponse({'success': True, 'count': len(selected_ids)})
            if action == 'load_selection':
                page_key = data.get('page_key', 'ajax_documentation_list')
                try:
                    selection = UserTableSelection.objects.get(user=request.user, page_key=page_key)
                    return JsonResponse({'success': True, 'selected_ids': selection.selected_ids})
                except UserTableSelection.DoesNotExist:
                    return JsonResponse({'success': True, 'selected_ids': []})

    if request.method == 'POST' and 'action' in request.POST:
        action = request.POST.get('action')
        selected_ids = request.POST.getlist('selected_ids')
        if not selected_ids:
            single_id = request.POST.get('id')
            if single_id:
                selected_ids = [single_id]
        if not selected_ids:
            try:
                selection = UserTableSelection.objects.get(user=request.user, page_key='ajax_documentation_list')
                selected_ids = selection.selected_ids or []
            except UserTableSelection.DoesNotExist:
                selected_ids = []
        if not selected_ids:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'error': 'Tidak ada item yang dipilih'}, status=400)
            messages.error(request, 'Tidak ada item yang dipilih')
            return redirect('manajemen_aplikasi:ajax_documentation_list')

        qs = ApiDocumentation.objects.filter(id__in=selected_ids, url__icontains='/ajax/').order_by('-id')
        if action == 'export_csv':
            return export_api_documentation_csv(qs)
        if action == 'export_excel':
            return export_api_documentation_excel(qs)
        if action in ('bulk_delete', 'delete_single'):
            deleted_count = qs.count()
            deleted_urls = list(qs.values_list('url', flat=True))
            if deleted_count:
                qs.delete()
                UserTableSelection.objects.filter(user=request.user, page_key='ajax_documentation_list').delete()
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': True, 'deleted': True, 'count': deleted_count, 'names': deleted_urls})
            messages.success(request, f'{deleted_count} endpoint berhasil dihapus')
            return redirect('manajemen_aplikasi:ajax_documentation_list')

    qs = ApiDocumentation.objects.filter(url__icontains='/ajax/')
    search_query = request.GET.get('search', '').strip()
    if search_query:
        qs = qs.filter(
            Q(url__icontains=search_query)
            | Q(method_type__icontains=search_query)
            | Q(description__icontains=search_query)
        )
    qs = qs.order_by('-id')

    total = qs.count()
    table = ApiDocumentationTable(qs)
    per_page = request.GET.get('per_page', '10')
    try:
        per_page = int(per_page)
        if per_page not in [10, 25, 50, 100]:
            per_page = 10
    except (ValueError, TypeError):
        per_page = 10
    RequestConfig(request, paginate={'per_page': per_page}).configure(table)

    context = {
        'table': table,
        'total': total,
        'search_query': search_query,
        'per_page': per_page,
    }

    is_htmx = request.headers.get('HX-Request') or request.headers.get('X-Requested-With') == 'XMLHttpRequest'
    if is_htmx:
        return render(request, 'manajemen_aplikasi/access/ma_ap_api_documentation/partials/_table.html', context)
    return render(request, 'manajemen_aplikasi/access/ma_ap_api_documentation/ajax_list.html', context)


@permission_required_403('pengaturan', 'api_documentation', 'view')
def api_documentation_detail(request, doc_id: int):
    doc: dict | None = None

    obj = ApiDocumentation.objects.filter(id=doc_id, is_active=True).values(
        'id', 'method_type', 'url', 'parameters', 'description', 'created_at', 'updated_at'
    ).first()
    if obj:
        obj['method_type_display'] = _normalize_method(obj.get('method_type'))
        obj['parameters_display'] = _normalize_parameters(obj.get('parameters'))
        doc = obj
    else:
        rows = _load_api_docs_from_markdown()
        for r in rows:
            if int(r.get('id') or 0) == int(doc_id):
                doc = r
                break

    context = {
        'doc': doc,
    }
    return render(request, 'manajemen_aplikasi/access/ma_ap_api_documentation/detail.html', context)


@permission_required_403('pengaturan', 'api_documentation', 'view')
def api_documentation_tester(request):
    endpoints = list(
        ApiDocumentation.objects.filter(is_active=True)
        .order_by('url', 'method_type')
        .values('id', 'method_type', 'url', 'parameters', 'description')
    )
    context = {
        'endpoints': endpoints,
    }
    return render(request, 'manajemen_aplikasi/access/ma_ap_api_documentation/tester.html', context)


@permission_required_403('pengaturan', 'api_documentation', 'create')
def api_documentation_create(request):
    if request.method == 'POST':
        method_type = (request.POST.get('method_type') or '').strip().upper()
        url = (request.POST.get('url') or '').strip()
        description = (request.POST.get('description') or '').strip()
        parameters_raw = (request.POST.get('parameters') or '').strip()
        is_active = request.POST.get('is_active') in ('1', 'true', 'True', 'on', 'yes')

        errors = []
        field_errors = {}
        if not method_type:
            errors.append('Method harus diisi!')
            field_errors['method_type'] = ['Method harus diisi!']
        if not url:
            errors.append('URL harus diisi!')
            field_errors['url'] = ['URL harus diisi!']
        if method_type and url and ApiDocumentation.objects.filter(method_type=method_type, url=url).exists():
            msg = 'Dokumentasi API untuk Method+URL tersebut sudah ada!'
            errors.append(msg)
            field_errors.setdefault('url', []).append(msg)

        params_val = None
        if parameters_raw:
            try:
                params_val = json.loads(parameters_raw)
            except Exception:
                params_val = parameters_raw

        is_htmx = request.headers.get('HX-Request') or request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        if errors:
            if is_htmx:
                resp = HttpResponse(status=204)
                resp['HX-Trigger'] = json.dumps({'api-doc-form-invalid': {'errors': errors, 'fieldErrors': field_errors}})
                return resp
            for e in errors:
                messages.error(request, e)
            return render(request, 'manajemen_aplikasi/access/ma_ap_api_documentation/form.html', {'errors_list': errors})

        ApiDocumentation.objects.create(
            method_type=method_type,
            url=url,
            description=description,
            parameters=params_val,
            is_active=is_active,
        )

        if is_htmx:
            resp = HttpResponse(status=204)
            resp['HX-Trigger'] = json.dumps({'api-doc-form-success': {'message': 'Dokumentasi API berhasil dibuat!', 'redirect': reverse('manajemen_aplikasi:api_documentation_list')}})
            return resp
        messages.success(request, '✅ Dokumentasi API berhasil dibuat!')
        return redirect('manajemen_aplikasi:api_documentation_list')

    return render(request, 'manajemen_aplikasi/access/ma_ap_api_documentation/form.html')


@permission_required_403('pengaturan', 'api_documentation', 'edit')
def api_documentation_edit(request, doc_id: int):
    obj = get_object_or_404(ApiDocumentation, id=doc_id)
    if request.method == 'POST':
        method_type = (request.POST.get('method_type') or '').strip().upper()
        url = (request.POST.get('url') or '').strip()
        description = (request.POST.get('description') or '').strip()
        parameters_raw = (request.POST.get('parameters') or '').strip()
        is_active = request.POST.get('is_active') in ('1', 'true', 'True', 'on', 'yes')

        errors = []
        field_errors = {}
        if not method_type:
            errors.append('Method harus diisi!')
            field_errors['method_type'] = ['Method harus diisi!']
        if not url:
            errors.append('URL harus diisi!')
            field_errors['url'] = ['URL harus diisi!']
        if method_type and url:
            if ApiDocumentation.objects.exclude(id=obj.id).filter(method_type=method_type, url=url).exists():
                msg = 'Dokumentasi API untuk Method+URL tersebut sudah ada!'
                errors.append(msg)
                field_errors.setdefault('url', []).append(msg)

        params_val = None
        if parameters_raw:
            try:
                params_val = json.loads(parameters_raw)
            except Exception:
                params_val = parameters_raw

        is_htmx = request.headers.get('HX-Request') or request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        if errors:
            if is_htmx:
                resp = HttpResponse(status=204)
                resp['HX-Trigger'] = json.dumps({'api-doc-form-invalid': {'errors': errors, 'fieldErrors': field_errors}})
                return resp
            for e in errors:
                messages.error(request, e)
            return render(request, 'manajemen_aplikasi/access/ma_ap_api_documentation/form.html', {'errors_list': errors, 'doc': obj, 'edit_mode': True})

        obj.method_type = method_type
        obj.url = url
        obj.description = description
        obj.parameters = params_val
        obj.is_active = is_active
        obj.save()

        if is_htmx:
            resp = HttpResponse(status=204)
            resp['HX-Trigger'] = json.dumps({'api-doc-form-success': {'message': 'Dokumentasi API berhasil diupdate!', 'redirect': reverse('manajemen_aplikasi:api_documentation_list')}})
            return resp
        messages.success(request, '✅ Dokumentasi API berhasil diupdate!')
        return redirect('manajemen_aplikasi:api_documentation_list')

    context = {
        'doc': obj,
        'edit_mode': True,
    }
    return render(request, 'manajemen_aplikasi/access/ma_ap_api_documentation/form.html', context)


@permission_required_403('pengaturan', 'api_documentation', 'delete')
def api_documentation_delete(request, doc_id: int):
    obj = get_object_or_404(ApiDocumentation, id=doc_id)
    if request.method == 'POST':
        url = obj.url
        obj.delete()
        messages.success(request, f'✅ Endpoint "{url}" berhasil dihapus!')
        return redirect('manajemen_aplikasi:api_documentation_list')
    return render(request, 'manajemen_aplikasi/access/ma_ap_api_documentation/delete.html', {'doc': obj})
