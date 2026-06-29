import random
import json
from datetime import date, datetime
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.urls import reverse
from django.utils import timezone
from django.db.models import Q, Count
from apps.manajemen.decorators import permission_required_403
from apps.manajemen.helpers import check_permission
from apps.api_simpeg.models import Pegawai
from .models import JenisSurvey, PeriodeSurvey, PertanyaanSurvey, RespondenSurvey, JawabanSurvey
from .models_pegawai_riwayat import PegawaiRiwayatData


def validate_nip(username):
    if username and len(username) == 18 and username.isdigit():
        return True
    return False


def _get_pegawai_by_nip(nip):
    return Pegawai.objects.filter(
        Q(nip_baru=nip) | Q(nip_lama=nip)
    ).first()


def _archive_pegawai_data(id_pegawai, tahun):
    existing = PegawaiRiwayatData.objects.filter(
        id_pegawai=id_pegawai,
        snapshot_at__year=tahun
    ).first()
    if existing:
        return
    pegawai = Pegawai.objects.filter(id_pegawai=id_pegawai).first()
    if not pegawai:
        return
    riwayat_data = {
        'id_pegawai': pegawai.id_pegawai,
        'nip_baru': pegawai.nip_baru,
        'nip_lama': pegawai.nip_lama,
        'namaPegawai': pegawai.nama_pegawai,
        'namaJabatan': pegawai.nama_jabatan,
        'nm_opd': pegawai.nm_opd,
        'nm_sub_opd': pegawai.nm_sub_opd,
        'namaGolongan': pegawai.nama_golongan,
        'namaPangkat': pegawai.nama_pangkat,
        'kodeEselon': pegawai.kode_eselon,
        'id_opd': pegawai.id_opd,
        'namaKategoriPegawai': pegawai.nama_kategori_pegawai,
    }
    jenis_survey = JenisSurvey.objects.filter(kode='SURVEY_360').first()
    if not jenis_survey:
        return
    periode = PeriodeSurvey.objects.filter(
        jenis_survey=jenis_survey,
        is_active=True,
        tanggal_mulai__lte=timezone.now(),
        tanggal_selesai__gte=timezone.now()
    ).first()
    if not periode:
        return
    PegawaiRiwayatData.objects.create(
        periode_survey=periode,
        jenis_survey=jenis_survey,
        id_pegawai=pegawai.id_pegawai,
        nip_baru=pegawai.nip_baru,
        nip_lama=pegawai.nip_lama,
        nama_pegawai=pegawai.nama_pegawai,
        nama_jabatan=pegawai.nama_jabatan,
        kode_eselon=pegawai.kode_eselon,
        id_opd=pegawai.id_opd,
        nm_opd=pegawai.nm_opd,
        nm_sub_opd=pegawai.nm_sub_opd,
        nama_golongan=pegawai.nama_golongan,
        nama_pangkat=pegawai.nama_pangkat,
        kategori_pegawai=pegawai.kategori_pegawai,
        nama_kategori_pegawai=pegawai.nama_kategori_pegawai,
        raw_data={},
    )


def _select_survey_target(pegawai_penilai, excluded_ids):
    current_year = date.today().year
    overload_ids = list(RespondenSurvey.objects.filter(
        statusData='submitted',
        created_at__year=current_year
    ).values('id_pegawaiDinilai').annotate(
        total=Count('id_pegawaiDinilai')
    ).filter(total__gte=6).values_list('id_pegawaiDinilai', flat=True))
    excluded = list(set(excluded_ids + list(overload_ids) + [pegawai_penilai.id_pegawai]))
    kode_eselon = str(pegawai_penilai.kode_eselon or '')

    def _select_by_opd():
        return Pegawai.objects.filter(
            id_opd=pegawai_penilai.id_opd
        ).exclude(id_pegawai__in=excluded).order_by('?').first()

    def _select_by_eselon():
        internal = random.randint(1, 100)
        if internal <= 10:
            target = Pegawai.objects.filter(
                kode_eselon=21
            ).exclude(id_pegawai__in=excluded).order_by('?').first()
            if target:
                return target
        target = Pegawai.objects.filter(
            kode_eselon=22
        ).exclude(id_pegawai__in=excluded).order_by('?').first()
        if target:
            return target
        target = Pegawai.objects.filter(
            kode_eselon=21
        ).exclude(id_pegawai__in=excluded).order_by('?').first()
        if target:
            return target
        return _select_by_opd()

    def _select_by_bawahan():
        if kode_eselon == '21':
            min_eselon = 22
        else:
            min_eselon = 23
        return Pegawai.objects.filter(
            id_opd=pegawai_penilai.id_opd,
            kode_eselon__gte=min_eselon,
        ).exclude(
            Q(kode_eselon=99) | Q(id_pegawai__in=excluded)
        ).order_by('?').first()

    def _select_by_atasan():
        return Pegawai.objects.filter(
            id_opd=pegawai_penilai.id_opd,
        ).exclude(id_pegawai__in=excluded).filter(
            Q(kode_eselon=21) | Q(kode_eselon=22) |
            Q(kode_eselon__startswith=1) | Q(kode_eselon__startswith=2)
        ).order_by('?').first()

    def _select_by_peer_level3():
        return Pegawai.objects.filter(
            id_opd=pegawai_penilai.id_opd,
            kode_eselon__startswith=3,
        ).exclude(
            Q(kode_eselon=99) | Q(id_pegawai__in=excluded)
        ).order_by('?').first()

    def _select_by_bawahan_level3():
        return Pegawai.objects.filter(
            id_opd=pegawai_penilai.id_opd,
        ).exclude(id_pegawai__in=excluded).filter(
            Q(kode_eselon__startswith=4) | Q(kode_eselon__startswith=5) |
            Q(kode_eselon=99)
        ).order_by('?').first()

    def _select_by_level2():
        return Pegawai.objects.filter(
            id_opd=pegawai_penilai.id_opd,
            kode_eselon__startswith=2,
        ).exclude(id_pegawai__in=excluded).order_by('?').first()

    def _select_by_level3_staff():
        return Pegawai.objects.filter(
            id_opd=pegawai_penilai.id_opd,
            kode_eselon__startswith=3,
        ).exclude(
            Q(kode_eselon=99) | Q(id_pegawai__in=excluded)
        ).order_by('?').first()

    def _select_by_level_other():
        return Pegawai.objects.filter(
            id_opd=pegawai_penilai.id_opd,
        ).exclude(id_pegawai__in=excluded).exclude(
            Q(kode_eselon__startswith=2) | Q(kode_eselon__startswith=3)
        ).order_by('?').first()

    if kode_eselon == '21':
        target = Pegawai.objects.filter(
            kode_eselon=22,
            id_opd=pegawai_penilai.id_opd,
        ).exclude(id_pegawai__in=excluded).order_by('?').first()
        if target:
            return target
        target = Pegawai.objects.filter(
            kode_eselon=22
        ).exclude(id_pegawai__in=excluded).order_by('?').first()
        if target:
            return target
        return _select_by_bawahan()

    if kode_eselon == '22':
        chance = random.randint(1, 100)
        result = None
        if chance <= 60:
            result = _select_by_eselon()
        elif chance <= 95:
            result = _select_by_bawahan()
        else:
            result = _select_by_opd()
        if not result:
            for fn in [_select_by_eselon, _select_by_bawahan, _select_by_opd]:
                result = fn()
                if result:
                    break
        return result

    first_digit = kode_eselon[:1] if kode_eselon else ''
    if first_digit == '3':
        chance = random.randint(1, 100)
        result = None
        if chance <= 30:
            result = _select_by_atasan()
        elif chance <= 60:
            result = _select_by_peer_level3()
        else:
            result = _select_by_bawahan_level3()
        if not result:
            for fn in [_select_by_atasan, _select_by_peer_level3, _select_by_bawahan_level3, _select_by_opd]:
                result = fn()
                if result:
                    break
        return result

    if first_digit == '4' or kode_eselon == '99':
        chance = random.randint(1, 100)
        result = None
        if chance <= 10:
            result = _select_by_level2()
        elif chance <= 50:
            result = _select_by_level3_staff()
        else:
            result = _select_by_level_other()
        if not result:
            for fn in [_select_by_level2, _select_by_level3_staff, _select_by_level_other, _select_by_opd]:
                result = fn()
                if result:
                    break
        return result

    return _select_by_opd()


def _determine_peran_penilai(pegawai_penilai, pegawai_dinilai):
    kode_penilai = str(pegawai_penilai.kode_eselon or '')
    kode_dinilai = str(pegawai_dinilai.kode_eselon or '')
    if kode_penilai in ('21', '22'):
        try:
            eselon_penilai = int(kode_penilai)
            eselon_dinilai = int(kode_dinilai)
            if eselon_penilai == eselon_dinilai:
                return '20'
            elif eselon_penilai > eselon_dinilai:
                return '10'
            else:
                return '30'
        except (ValueError, TypeError):
            pass
    digit_penilai = int(kode_penilai[:1]) if kode_penilai and kode_penilai[:1].isdigit() else 0
    digit_dinilai = int(kode_dinilai[:1]) if kode_dinilai and kode_dinilai[:1].isdigit() else 0
    if digit_penilai == digit_dinilai:
        return '20'
    elif digit_penilai < digit_dinilai:
        return '30'
    else:
        return '10'


@permission_required_403('survey360', 'survey360_fill', 'view')
def survey360_index(request):
    nip = request.user.email or request.user.username
    if not validate_nip(nip):
        return redirect('survey:survey360_invalid_nip')
    pegawai_penilai = _get_pegawai_by_nip(nip)
    if not pegawai_penilai:
        messages.error(request, 'Data pegawai tidak ditemukan.')
        return redirect('dashboard:index')

    respondens = RespondenSurvey.objects.filter(
        id_pegawaiPenilai=pegawai_penilai.id_pegawai,
        jenis_survey__kode='SURVEY_360'
    ).select_related('jenis_survey').order_by('-created_at')

    survey_list = []
    for r in respondens:
        pegawai_dinilai = Pegawai.objects.filter(id_pegawai=r.id_pegawaiDinilai).first()
        survey_list.append({
            'id': r.id,
            'nip_pegawaiDinilai': r.nip_pegawaiDinilai,
            'nama_pegawaiDinilai': pegawai_dinilai.nama_pegawai if pegawai_dinilai else '-',
            'tahun': r.created_at.year,
            'tanggal_submit': r.updated_at.strftime('%d/%m/%Y %H:%M') if r.statusData == 'submitted' else '',
            'statusData': r.statusData,
            'peranPenilai': r.peranPenilai,
        })

    context = {
        'surveys': survey_list,
        'pegawai_penilai': pegawai_penilai,
        'page_title': 'Survey 360',
    }
    return render(request, 'survey360/index.html', context)


@permission_required_403('survey360', 'survey360_fill', 'view')
def survey360_invalid_nip(request):
    return render(request, 'survey360/invalid_nip.html', {'page_title': 'NIP Tidak Valid'})


@permission_required_403('survey360', 'survey360_fill', 'view')
def survey360_preview_target(request):
    nip = request.user.email or request.user.username
    if not validate_nip(nip):
        return JsonResponse({
            'success': False,
            'message': 'Username tidak valid untuk Survey 360. Harus berupa NIP 18 digit.',
            'redirect': reverse('survey:survey360_invalid_nip')
        })
    pegawai_penilai = _get_pegawai_by_nip(nip)
    if not pegawai_penilai:
        return JsonResponse({'success': False, 'message': 'Data pegawai tidak ditemukan'})

    current_year = date.today().year
    jenis_survey = JenisSurvey.objects.filter(kode='SURVEY_360').first()
    if not jenis_survey:
        return JsonResponse({'success': False, 'message': 'Jenis Survey 360 belum dikonfigurasi.'})

    existing_draft = RespondenSurvey.objects.filter(
        id_pegawaiPenilai=pegawai_penilai.id_pegawai,
        statusData='draft',
        created_at__year=current_year,
        jenis_survey=jenis_survey,
    ).first()

    if existing_draft:
        pegawai_dinilai = Pegawai.objects.filter(id_pegawai=existing_draft.id_pegawaiDinilai).first()
        if pegawai_dinilai:
            return JsonResponse({
                'success': True,
                'isDraft': True,
                'message': 'Anda memiliki survey yang belum diselesaikan',
                'targetInfo': {
                    'id_pegawai': pegawai_dinilai.id_pegawai,
                    'nama': pegawai_dinilai.nama_pegawai,
                    'nip': pegawai_dinilai.nip_baru or pegawai_dinilai.nip_lama or '',
                    'jabatan': pegawai_dinilai.nama_jabatan or '-',
                    'opd': pegawai_dinilai.nm_opd or '-',
                    'created_at': existing_draft.created_at.strftime('%d/%m/%Y %H:%M'),
                }
            })
    already_rated = list(RespondenSurvey.objects.filter(
        id_pegawaiPenilai=pegawai_penilai.id_pegawai,
        created_at__year=current_year,
        jenis_survey=jenis_survey,
    ).values_list('id_pegawaiDinilai', flat=True))

    target = _select_survey_target(pegawai_penilai, already_rated)
    if not target:
        return JsonResponse({
            'success': False,
            'message': 'Semua pegawai di unit kerja Anda sudah dinilai untuk tahun {}. Silakan tunggu tahun depan.'.format(current_year)
        })

    peran = _determine_peran_penilai(pegawai_penilai, target)
    _archive_pegawai_data(pegawai_penilai.id_pegawai, current_year)
    _archive_pegawai_data(target.id_pegawai, current_year)

    responden = RespondenSurvey.objects.create(
        jenis_survey=jenis_survey,
        id_pegawaiPenilai=pegawai_penilai.id_pegawai,
        nip_pegawaiPenilai=pegawai_penilai.nip_baru or pegawai_penilai.nip_lama or '',
        id_pegawaiDinilai=target.id_pegawai,
        nip_pegawaiDinilai=target.nip_baru or target.nip_lama or '',
        peranPenilai=peran,
        statusData='draft',
    )

    return JsonResponse({
        'success': True,
        'isDraft': False,
        'message': 'Target survey berhasil ditemukan dan terkunci',
        'targetInfo': {
            'id_pegawai': target.id_pegawai,
            'nama': target.nama_pegawai,
            'nip': target.nip_baru or target.nip_lama or '',
            'jabatan': target.nama_jabatan or '-',
            'opd': target.nm_opd or '-',
            'created_at': responden.created_at.strftime('%d/%m/%Y %H:%M'),
        }
    })


@permission_required_403('survey360', 'survey360_fill', 'create')
def survey360_create(request):
    nip = request.user.email or request.user.username
    if not validate_nip(nip):
        return redirect('survey:survey360_invalid_nip')

    pegawai_penilai = _get_pegawai_by_nip(nip)
    if not pegawai_penilai:
        messages.error(request, 'Data pegawai tidak ditemukan.')
        return redirect('survey:survey360_index')

    current_year = date.today().year
    jenis_survey = JenisSurvey.objects.filter(kode='SURVEY_360').first()
    if not jenis_survey:
        messages.error(request, 'Jenis Survey 360 belum dikonfigurasi.')
        return redirect('survey:survey360_index')

    existing = RespondenSurvey.objects.filter(
        id_pegawaiPenilai=pegawai_penilai.id_pegawai,
        statusData='draft',
        created_at__year=current_year,
        jenis_survey=jenis_survey,
    ).first()

    if existing:
        pegawai_dinilai = Pegawai.objects.filter(id_pegawai=existing.id_pegawaiDinilai).first()
        peran = existing.peranPenilai
        tahun = existing.created_at.year
        survey_id = existing.id
    else:
        already = list(RespondenSurvey.objects.filter(
            id_pegawaiPenilai=pegawai_penilai.id_pegawai,
            created_at__year=current_year,
            jenis_survey=jenis_survey,
        ).values_list('id_pegawaiDinilai', flat=True))
        target = _select_survey_target(pegawai_penilai, already)
        if not target:
            messages.warning(request, 'Tidak ada pegawai yang dapat disurvey saat ini.')
            return redirect('survey:survey360_index')

        peran = _determine_peran_penilai(pegawai_penilai, target)
        _archive_pegawai_data(pegawai_penilai.id_pegawai, current_year)
        _archive_pegawai_data(target.id_pegawai, current_year)

        existing = RespondenSurvey.objects.create(
            jenis_survey=jenis_survey,
            id_pegawaiPenilai=pegawai_penilai.id_pegawai,
            nip_pegawaiPenilai=pegawai_penilai.nip_baru or pegawai_penilai.nip_lama or '',
            id_pegawaiDinilai=target.id_pegawai,
            nip_pegawaiDinilai=target.nip_baru or target.nip_lama or '',
            peranPenilai=peran,
            statusData='draft',
        )
        pegawai_dinilai = target
        tahun = existing.created_at.year
        survey_id = existing.id

    pertanyaan_list = PertanyaanSurvey.objects.filter(
        jenis_survey=jenis_survey,
        is_active=True
    ).order_by('urutan')

    context = {
        'pegawai_penilai': pegawai_penilai,
        'pegawai_dinilai': pegawai_dinilai,
        'peran_penilai': peran,
        'tahun_penilaian': tahun,
        'survey_id': survey_id,
        'pertanyaan_list': pertanyaan_list,
        'page_title': 'Survey 360 - Penilaian',
        'existing_jawaban': {},
    }
    if existing and existing.jawaban.exists():
        for j in existing.jawaban.all():
            context['existing_jawaban'][j.pertanyaan_id] = j.nilai

    return render(request, 'survey360/create.html', context)


@permission_required_403('survey360', 'survey360_fill', 'submit')
def survey360_store(request):
    if request.method != 'POST':
        return redirect('survey:survey360_index')

    survey_id = request.POST.get('survey_id')
    id_pegawai_penilai = request.POST.get('id_pegawaiPenilai')
    id_pegawai_dinilai = request.POST.get('id_pegawaiDinilai')
    tahun = request.POST.get('tahunPenilaian', str(date.today().year))

    if not survey_id:
        messages.error(request, 'Survey tidak ditemukan.')
        return redirect('survey:survey360_index')

    responden = get_object_or_404(RespondenSurvey, id=survey_id)
    if responden.statusData == 'submitted':
        messages.warning(request, 'Survey ini sudah pernah disubmit.')
        return redirect('survey:survey360_index')

    jenis_survey = JenisSurvey.objects.filter(kode='SURVEY_360').first()
    pertanyaan_list = PertanyaanSurvey.objects.filter(
        jenis_survey=jenis_survey,
        is_active=True
    ).order_by('urutan')

    errors = []
    jawaban_data = {}
    for pt in pertanyaan_list:
        key = f'pertanyaan_{pt.id}'
        val = request.POST.get(key)
        if not val:
            errors.append(f'{pt.judul} harus diisi.')
            continue
        try:
            nilai = int(val)
            if nilai < 1 or nilai > 5:
                errors.append(f'{pt.judul} harus bernilai 1-5.')
                continue
            jawaban_data[pt.id] = nilai
        except (ValueError, TypeError):
            errors.append(f'{pt.judul} tidak valid.')

    if errors:
        return JsonResponse({'success': False, 'errors': errors})

    _archive_pegawai_data(id_pegawai_penilai, tahun)
    _archive_pegawai_data(id_pegawai_dinilai, tahun)

    JawabanSurvey.objects.filter(responden=responden).delete()
    for pt_id, nilai in jawaban_data.items():
        JawabanSurvey.objects.create(
            responden=responden,
            pertanyaan_id=pt_id,
            nilai=nilai,
        )

    responden.statusData = 'submitted'
    responden.save()

    pegawai_dinilai = Pegawai.objects.filter(id_pegawai=responden.id_pegawaiDinilai).first()
    nama = pegawai_dinilai.nama_pegawai if pegawai_dinilai else 'Pegawai'

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({
            'success': True,
            'message': f'Survey 360° untuk {nama} berhasil disubmit!',
            'pegawai_dinilai': nama,
        })

    messages.success(request, f'Survey 360° untuk {nama} berhasil disubmit!')
    return redirect('survey:survey360_index')


@permission_required_403('survey360', 'survey360_report', 'view')
def survey360_laporan(request):
    from collections import defaultdict
    import json
    current_year = date.today().year
    respondens = RespondenSurvey.objects.filter(
        jenis_survey__kode='SURVEY_360'
    ).select_related('jenis_survey').order_by('-updated_at').prefetch_related('jawaban')

    total_survey = respondens.count()
    tahun_ini = respondens.filter(updated_at__year=current_year).count()

    nips = list(respondens.values_list('nip_pegawaiDinilai', flat=True).distinct())
    pegawai_map = {}
    if nips:
        pegawais = Pegawai.objects.filter(
            Q(nip_baru__in=nips) | Q(nip_lama__in=nips)
        ).values('nip_baru', 'nip_lama', 'nm_opd', 'nama_pegawai')
        for p in pegawais:
            key = p['nip_baru'] or p['nip_lama'] or ''
            if key:
                pegawai_map[key] = {'nm_opd': p['nm_opd'] or '-', 'nama_pegawai': p['nama_pegawai'] or '-'}

    opd_groups = defaultdict(lambda: {'total': 0, 'submitted': 0, 'draft': 0, 'respondens': []})
    for r in respondens:
        nip = r.nip_pegawaiDinilai
        info = pegawai_map.get(nip, {'nm_opd': 'Unit Tidak Diketahui', 'nama_pegawai': '-'})
        nm_opd = info['nm_opd']
        opd_groups[nm_opd]['total'] += 1
        if r.statusData == 'submitted':
            opd_groups[nm_opd]['submitted'] += 1
        else:
            opd_groups[nm_opd]['draft'] += 1
        updated_str = r.updated_at.strftime('%d/%m/%Y %H:%M') if r.updated_at else '-'
        opd_groups[nm_opd]['respondens'].append({
            'id': r.id,
            'nip_pegawaiDinilai': r.nip_pegawaiDinilai or '-',
            'nama_pegawai': info['nama_pegawai'],
            'peranPenilai': r.peranPenilai or '',
            'statusData': r.statusData,
            'updated_at': updated_str,
        })

    unit_kerja_list = [{'nama': opd, **data} for opd, data in sorted(opd_groups.items())]

    context = {
        'respondens': respondens,
        'total_survey': total_survey,
        'tahun_ini': tahun_ini,
        'unit_kerja_list': unit_kerja_list,
        'unit_kerja_json': json.dumps(unit_kerja_list),
        'current_year': current_year,
        'page_title': 'Laporan Survey 360',
    }
    return render(request, 'survey360/laporan.html', context)


@permission_required_403('survey360', 'survey360_report', 'export')
def survey360_download_excel(request):
    try:
        tahun = request.GET.get('tahun')
        qs = RespondenSurvey.objects.filter(
            statusData='submitted',
            jenis_survey__kode='SURVEY_360'
        ).select_related('jenis_survey')

        if tahun:
            qs = qs.filter(updated_at__year=int(tahun))

        qs = qs.order_by('nip_pegawaiDinilai')

        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Survey 360'

        headers = ['NIP Dinilai', 'Peran Penilai',
                   'Survey 01', 'Survey 02', 'Survey 03', 'Survey 04',
                   'Survey 05', 'Survey 06', 'Survey 07']
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        row = 2
        for r in qs:
            jawaban = {j.pertanyaan.kode_pertanyaan: j.nilai for j in r.jawaban.select_related('pertanyaan').all()}
            ws.cell(row=row, column=1, value=f"'{r.nip_pegawaiDinilai}")
            ws.cell(row=row, column=2, value=r.peranPenilai)
            ws.cell(row=row, column=3, value=jawaban.get('survey01', ''))
            ws.cell(row=row, column=4, value=jawaban.get('survey02', ''))
            ws.cell(row=row, column=5, value=jawaban.get('survey03', ''))
            ws.cell(row=row, column=6, value=jawaban.get('survey04', ''))
            ws.cell(row=row, column=7, value=jawaban.get('survey05', ''))
            ws.cell(row=row, column=8, value=jawaban.get('survey06', ''))
            ws.cell(row=row, column=9, value=jawaban.get('survey07', ''))
            row += 1

        filename = f'Survey360_Export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response

    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Gagal mengexport data: {str(e)}'
        }, status=500)


@permission_required_403('survey360', 'survey360_report', 'reset')
def survey360_reset(request, pk):
    obj = get_object_or_404(RespondenSurvey, id=pk, jenis_survey__kode='SURVEY_360')
    if request.method == 'POST':
        JawabanSurvey.objects.filter(responden=obj).delete()
        obj.statusData = 'draft'
        obj.save()
        messages.success(request, 'Survey 360 berhasil direset ke draft.')
        return redirect('survey:survey360_laporan')
    context = {
        'obj': obj,
        'page_title': 'Reset Survey 360',
    }
    return render(request, 'survey360/reset_confirm.html', context)
